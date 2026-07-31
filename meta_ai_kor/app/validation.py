from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.excel import RESULT_HEADERS
from app.glossary import MappingGlossary, normalize_header
from app.models import (
    ColumnResult,
    SourceRow,
    ValidationIssue,
    ValidationReport,
)
from app.normalization import is_valid_korean_name


def validate_results(
    sources: Iterable[SourceRow],
    results: Iterable[ColumnResult],
    glossary: MappingGlossary,
    *,
    auto_confirm_threshold: int = 85,
) -> ValidationReport:
    source_list = list(sources)
    result_list = list(results)
    issues: list[ValidationIssue] = []
    if len(source_list) != len(result_list):
        issues.append(
            ValidationIssue(
                code="row_count_mismatch",
                severity="error",
                message=(
                    f"입력 {len(source_list)}행과 결과 "
                    f"{len(result_list)}행이 다릅니다."
                ),
            )
        )
    result_by_id = {result.source_id: result for result in result_list}
    if len(result_by_id) != len(result_list):
        issues.append(
            ValidationIssue(
                code="duplicate_source_id",
                severity="error",
                message="결과 source_id가 중복되었습니다.",
            )
        )
    for index, source in enumerate(source_list):
        result = result_by_id.get(source.source_id)
        if result is None:
            issues.append(
                _issue(
                    source.source_id,
                    "missing_result",
                    "error",
                    "원본행에 대응하는 결과가 없습니다.",
                )
            )
            continue
        if index < len(result_list) and result_list[index].source_id != source.source_id:
            issues.append(
                _issue(
                    source.source_id,
                    "row_order_changed",
                    "error",
                    "결과 행 순서가 원본과 다릅니다.",
                )
            )
        if not result.english_full_name.strip():
            issues.append(
                _issue(
                    source.source_id,
                    "empty_full_name",
                    "error",
                    "영문 Full Name이 비어 있습니다.",
                )
            )
        if not result.korean_attribute_name.strip():
            issues.append(
                _issue(
                    source.source_id,
                    "empty_korean_name",
                    "error",
                    "한글속성명이 비어 있습니다.",
                )
            )
        elif not is_valid_korean_name(result.korean_attribute_name):
            issues.append(
                _issue(
                    source.source_id,
                    "invalid_korean_name_format",
                    "error",
                    "한글속성명에는 한글과 숫자만 사용할 수 있습니다.",
                    {"actual": result.korean_attribute_name},
                )
            )
        if any(
            placeholder in result.korean_attribute_name
            for placeholder in ("미정", "불명", "알수없음")
        ):
            issues.append(
                _issue(
                    source.source_id,
                    "placeholder_korean_name",
                    "error",
                    "한글속성명에 자리표시자가 남아 있습니다.",
                    {"actual": result.korean_attribute_name},
                )
            )
        compact_column = re.sub(
            r"[^A-Z0-9]",
            "",
            source.column_name.upper(),
        )
        compact_components = "".join(
            component.source_fragment for component in result.components
        )
        if compact_components != compact_column:
            issues.append(
                _issue(
                    source.source_id,
                    "component_coverage_mismatch",
                    "error",
                    "컴포넌트 연결이 원본 컬럼명을 복원하지 못합니다.",
                    {
                        "expected": compact_column,
                        "actual": compact_components,
                    },
                )
            )
        for component in result.components:
            if not component.full_name or not component.korean_word:
                issues.append(
                    _issue(
                        source.source_id,
                        "empty_component_meaning",
                        "error",
                        "컴포넌트 Full Name 또는 한글단어가 비어 있습니다.",
                        {"component": component.model_dump()},
                    )
                )
            if component.origin == "mapping":
                exact = any(
                    entry.full_name == component.full_name
                    and entry.korean_word == component.korean_word
                    for entry in glossary.entries_for(
                        component.source_fragment
                    )
                )
                if not exact:
                    issues.append(
                        _issue(
                            source.source_id,
                            "mapping_evidence_mismatch",
                            "error",
                            "mapping 출처 컴포넌트가 사전과 일치하지 않습니다.",
                            {"component": component.model_dump()},
                        )
                    )
            evidence_token = (
                f"{component.source_fragment}→"
                f"{component.full_name}→{component.korean_word}"
            )
            if evidence_token not in result.evidence:
                issues.append(
                    _issue(
                        source.source_id,
                        "evidence_mismatch",
                        "error",
                        "변환근거가 실제 컴포넌트와 일치하지 않습니다.",
                        {"expected_token": evidence_token},
                    )
                )
        if not 0 <= result.confidence <= 100:
            issues.append(
                _issue(
                    source.source_id,
                    "confidence_out_of_range",
                    "error",
                    "신뢰도는 0~100이어야 합니다.",
                )
            )
        elif result.confidence < auto_confirm_threshold:
            issues.append(
                _issue(
                    source.source_id,
                    "low_confidence",
                    "warning",
                    "자동확정 임계값보다 신뢰도가 낮습니다.",
                    {
                        "actual": result.confidence,
                        "threshold": auto_confirm_threshold,
                    },
                )
            )

    contexts: dict[
        tuple[str, str, str, str],
        set[tuple[str, str]],
    ] = defaultdict(set)
    source_by_id = {source.source_id: source for source in source_list}
    for result in result_list:
        source = source_by_id.get(result.source_id)
        if source is not None:
            contexts[source.context_key].add(
                (
                    result.english_full_name,
                    result.korean_attribute_name,
                )
            )
    for context_key, values in contexts.items():
        if len(values) > 1:
            issues.append(
                ValidationIssue(
                    code="inconsistent_context_result",
                    severity="error",
                    message="동일 문맥 키에서 서로 다른 결과가 생성되었습니다.",
                    details={
                        "context_key": list(context_key),
                        "values": [list(value) for value in sorted(values)],
                    },
                )
            )
    error_count = sum(issue.severity == "error" for issue in issues)
    warning_count = sum(issue.severity == "warning" for issue in issues)
    affected_source_ids = {
        issue.source_id for issue in issues if issue.source_id is not None
    }
    return ValidationReport(
        is_valid=error_count == 0,
        issues=issues,
        stats={
            "source_count": len(source_list),
            "result_count": len(result_list),
            "error_count": error_count,
            "warning_count": warning_count,
            "affected_source_count": len(affected_source_ids),
        },
    )


def apply_validation_status(
    results: Iterable[ColumnResult],
    report: ValidationReport,
    *,
    auto_confirm_threshold: int = 85,
) -> list[ColumnResult]:
    errors_by_id: dict[str, list[str]] = defaultdict(list)
    warnings_by_id: dict[str, list[str]] = defaultdict(list)
    for issue in report.issues:
        if issue.source_id is None:
            continue
        target = (
            errors_by_id if issue.severity == "error" else warnings_by_id
        )
        target[issue.source_id].append(issue.code)
    output: list[ColumnResult] = []
    for result in results:
        errors = errors_by_id.get(result.source_id, [])
        warnings = warnings_by_id.get(result.source_id, [])
        inferred = any(
            component.origin == "inference" for component in result.components
        )
        if errors:
            status = "검증실패"
        elif (
            result.confidence >= auto_confirm_threshold
            and not inferred
            and result.review_stratum == "deterministic"
        ):
            status = "자동확정"
        else:
            status = "검토필요"
        output.append(
            result.model_copy(
                update={
                    "status": status,
                    "validation_codes": [*errors, *warnings],
                }
            )
        )
    return output


def validate_output_workbook(
    input_path: str | Path,
    output_path: str | Path,
) -> ValidationReport:
    input_workbook = load_workbook(
        input_path,
        read_only=True,
        data_only=True,
    )
    output_workbook = load_workbook(
        output_path,
        read_only=True,
        data_only=True,
    )
    input_sheet = input_workbook.active
    output_sheet = output_workbook["한글속성명_결과"]
    issues: list[ValidationIssue] = []
    if input_sheet.max_row != output_sheet.max_row:
        issues.append(
            ValidationIssue(
                code="output_row_count_mismatch",
                severity="error",
                message="출력 XLSX 행 수가 원본과 다릅니다.",
            )
        )
    max_row = min(input_sheet.max_row, output_sheet.max_row)
    changed_cells = 0
    for row in range(1, max_row + 1):
        for column in range(1, 13):
            before = _blank_normalized(input_sheet.cell(row, column).value)
            after = _blank_normalized(output_sheet.cell(row, column).value)
            if before != after:
                changed_cells += 1
    if changed_cells:
        issues.append(
            ValidationIssue(
                code="original_cell_changed",
                severity="error",
                message=f"원본 12개 컬럼의 값 {changed_cells}개가 변경되었습니다.",
                details={"changed_cell_count": changed_cells},
            )
        )
    actual_headers = [
        normalize_header(output_sheet.cell(1, column).value)
        for column in range(13, 18)
    ]
    expected_headers = [normalize_header(value) for value in RESULT_HEADERS]
    if actual_headers != expected_headers:
        issues.append(
            ValidationIssue(
                code="result_headers_mismatch",
                severity="error",
                message="결과 5개 컬럼 헤더가 계약과 다릅니다.",
                details={
                    "expected": expected_headers,
                    "actual": actual_headers,
                },
            )
        )
    input_workbook.close()
    output_workbook.close()
    error_count = sum(issue.severity == "error" for issue in issues)
    return ValidationReport(
        is_valid=error_count == 0,
        issues=issues,
        stats={
            "error_count": error_count,
            "warning_count": 0,
            "changed_original_cell_count": changed_cells,
        },
    )


def _issue(
    source_id: str,
    code: str,
    severity: str,
    message: str,
    details: dict[str, object] | None = None,
) -> ValidationIssue:
    return ValidationIssue(
        source_id=source_id,
        code=code,
        severity=severity,  # type: ignore[arg-type]
        message=message,
        details=details or {},
    )


def _blank_normalized(value):
    return None if value == "" else value

