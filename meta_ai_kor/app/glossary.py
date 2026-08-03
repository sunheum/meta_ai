from __future__ import annotations

import math
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.exceptions import InputWorkbookError
from app.models import MappingEntry, SourceRow


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s*\(\*\)\s*$", "", text.strip())
    return re.sub(r"\s+", "", text)


@dataclass(frozen=True, slots=True)
class MappingGlossary:
    by_abbreviation: dict[str, tuple[MappingEntry, ...]]
    by_abbreviation_full_name: dict[
        tuple[str, str], tuple[MappingEntry, ...]
    ]
    by_abbreviation_korean: dict[
        tuple[str, str], tuple[MappingEntry, ...]
    ]
    ambiguous_korean: frozenset[str] = field(default_factory=frozenset)
    ambiguous_full_name: frozenset[str] = field(default_factory=frozenset)
    source_path: str | None = None

    @classmethod
    def from_entries(
        cls,
        entries: Iterable[MappingEntry],
        source_path: str | None = None,
    ) -> "MappingGlossary":
        by_abbreviation: dict[str, list[MappingEntry]] = defaultdict(list)
        by_full_name: dict[tuple[str, str], list[MappingEntry]] = defaultdict(
            list
        )
        by_korean: dict[tuple[str, str], list[MappingEntry]] = defaultdict(list)
        seen: set[tuple[str, str, str, int]] = set()
        for entry in entries:
            key = (
                entry.abbreviation,
                entry.full_name,
                entry.korean_word,
                entry.occurrence_count,
            )
            if key in seen:
                continue
            seen.add(key)
            by_abbreviation[entry.abbreviation].append(entry)
            by_full_name[(entry.abbreviation, entry.full_name)].append(entry)
            by_korean[(entry.abbreviation, entry.korean_word)].append(entry)

        def freeze(
            values: dict[object, list[MappingEntry]],
        ) -> dict[object, tuple[MappingEntry, ...]]:
            return {
                key: tuple(
                    sorted(
                        items,
                        key=lambda item: (
                            -item.occurrence_count,
                            item.full_name,
                            item.korean_word,
                        ),
                    )
                )
                for key, items in values.items()
            }

        frozen_by_abbreviation = freeze(by_abbreviation)
        ambiguous_korean = frozenset(
            abbreviation
            for abbreviation, items in frozen_by_abbreviation.items()
            if len({item.korean_word for item in items}) > 1
        )
        ambiguous_full_name = frozenset(
            abbreviation
            for abbreviation, items in frozen_by_abbreviation.items()
            if len({item.full_name for item in items}) > 1
        )
        return cls(
            by_abbreviation=frozen_by_abbreviation,  # type: ignore[arg-type]
            by_abbreviation_full_name=freeze(by_full_name),  # type: ignore[arg-type]
            by_abbreviation_korean=freeze(by_korean),  # type: ignore[arg-type]
            ambiguous_korean=ambiguous_korean,
            ambiguous_full_name=ambiguous_full_name,
            source_path=source_path,
        )

    @classmethod
    def from_xlsx(cls, path: str | Path) -> "MappingGlossary":
        source_path = Path(path)
        try:
            workbook = load_workbook(
                source_path,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise InputWorkbookError(
                f"매핑 엑셀 파일을 열 수 없습니다: {exc}"
            ) from exc
        worksheet = workbook.active
        header_row, indexes = _find_headers(
            worksheet,
            {"영문약어", "영문FullName", "한글단어"},
        )
        entries: list[MappingEntry] = []
        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            abbreviation = _cell_text(row, indexes["영문약어"])
            full_name = _cell_text(row, indexes["영문FullName"])
            korean_word = _cell_text(row, indexes["한글단어"])
            if not abbreviation or not full_name or not korean_word:
                continue
            entries.append(
                MappingEntry(
                    abbreviation=abbreviation,
                    full_name=full_name,
                    korean_word=korean_word,
                    occurrence_count=_cell_int(
                        row,
                        indexes.get("출현건수"),
                    ),
                )
            )
        workbook.close()
        if not entries:
            raise InputWorkbookError("매핑 사전에 유효한 행이 없습니다.")
        return cls.from_entries(entries, source_path=str(source_path))

    def entries_for(self, abbreviation: str) -> tuple[MappingEntry, ...]:
        return self.by_abbreviation.get(abbreviation.upper(), ())

    def contains(self, abbreviation: str) -> bool:
        return abbreviation.upper() in self.by_abbreviation

    def resolve(
        self,
        abbreviation: str,
        source: SourceRow,
    ) -> tuple[MappingEntry | None, bool]:
        entries = self.entries_for(abbreviation)
        if not entries:
            return None, False
        ranked = sorted(
            entries,
            key=lambda entry: (
                -self.context_score(entry, source),
                -entry.occurrence_count,
                entry.full_name,
                entry.korean_word,
            ),
        )
        top = ranked[0]
        top_score = self.context_score(top, source)
        tied_meanings = {
            (entry.full_name, entry.korean_word)
            for entry in ranked
            if self.context_score(entry, source) == top_score
            and entry.occurrence_count == top.occurrence_count
        }
        return top, len(tied_meanings) > 1 or abbreviation.upper() in (
            self.ambiguous_korean | self.ambiguous_full_name
        )

    def context_score(self, entry: MappingEntry, source: SourceRow) -> float:
        description = _normalize_korean(source.table_description)
        korean_word = _normalize_korean(entry.korean_word)
        score = math.log1p(entry.occurrence_count)
        if korean_word and korean_word in description:
            score += 8.0
        full_name_tokens = {
            token for token in re.split(r"[^A-Z0-9]+", entry.full_name) if token
        }
        table_tokens = {
            token
            for token in re.split(r"[^A-Z0-9]+", source.table_name.upper())
            if token
        }
        if full_name_tokens & table_tokens:
            score += 2.0
        return round(score, 6)

    def __len__(self) -> int:
        return sum(len(items) for items in self.by_abbreviation.values())


def _find_headers(
    worksheet,
    wanted: set[str],
) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(30, worksheet.max_row),
            values_only=True,
        ),
        start=1,
    ):
        indexes = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None
        }
        if wanted.issubset(indexes):
            return row_number, indexes
    raise InputWorkbookError(
        f"필수 헤더를 찾지 못했습니다: {', '.join(sorted(wanted))}"
    )


def _cell_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _cell_int(row: tuple[object, ...], index: int | None) -> int:
    if index is None or index >= len(row) or row[index] is None:
        return 0
    try:
        return max(0, int(row[index]))
    except (TypeError, ValueError):
        return 0


def _normalize_korean(value: str) -> str:
    return re.sub(r"[\s_\-/·(),.]", "", value)

