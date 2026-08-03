from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.excel import read_source_workbook, write_result_workbook
from app.glossary import MappingGlossary
from app.llm import NamingModel
from app.models import (
    ColumnResult,
    NameComponent,
    WorkflowOptions,
)
from app.workflow import (
    NamingWorkflow,
    assign_review_strata,
)

_EVIDENCE_COMPONENT = re.compile(
    r"(?P<fragment>[^→|]+)→(?P<full>[^→|]+)→"
    r"(?P<korean>.*?)\[(?P<origin>mapping|inference|numeric)\]"
)


def read_existing_results(
    result_path: str | Path,
) -> list[ColumnResult]:
    workbook = load_workbook(result_path, read_only=True, data_only=True)
    worksheet = workbook["한글속성명_결과"]
    results: list[ColumnResult] = []
    for excel_row, values in enumerate(
        worksheet.iter_rows(
            min_row=2,
            min_col=5,
            max_col=17,
            values_only=True,
        ),
        start=2,
    ):
        column_name = str(values[0] or "").strip().upper()
        full_name = str(values[8] or "").strip()
        korean_name = str(values[9] or "").strip()
        status = str(values[10] or "").strip()
        confidence = int(values[11] or 0)
        evidence = str(values[12] or "").strip()
        components = _parse_evidence(evidence)
        if not components:
            compact = re.sub(r"[^A-Z0-9]", "", column_name)
            components = [
                NameComponent(
                    source_fragment=compact,
                    full_name=full_name or compact,
                    korean_word=korean_name or "미정",
                    origin="inference",
                    start=0,
                    end=len(compact),
                )
            ]
        inferred = any(
            component.origin == "inference" for component in components
        )
        results.append(
            ColumnResult(
                source_id=f"row-{excel_row}",
                components=components,
                english_full_name=full_name,
                korean_attribute_name=korean_name,
                status=status,
                confidence=confidence,
                evidence=evidence,
                reason="기존 결과 복원",
                review_stratum=(
                    "unmapped_inference"
                    if inferred
                    else (
                        "deterministic"
                        if status == "자동확정"
                        else "mapping_ambiguity"
                    )
                ),
            )
        )
    workbook.close()
    return results


async def repair_failed_workbook(
    *,
    source_path: str | Path,
    result_path: str | Path,
    mapping_path: str | Path,
    output_path: str | Path,
    model: NamingModel,
    options: WorkflowOptions,
) -> tuple[list, list[ColumnResult]]:
    sources = read_source_workbook(source_path)
    existing = read_existing_results(result_path)
    if [source.source_id for source in sources] != [
        result.source_id for result in existing
    ]:
        raise ValueError("원본과 기존 결과의 source_id 순서가 다릅니다.")
    glossary = MappingGlossary.from_xlsx(mapping_path)
    workflow = NamingWorkflow(glossary, model)
    source_by_id = {source.source_id: source for source in sources}
    failed_ids = {
        result.source_id
        for result in existing
        if result.status == "검증실패"
    }
    requests = [
        workflow._resolution_request(source_by_id[source_id])
        for source_id in sorted(
            failed_ids,
            key=lambda value: int(value.split("-")[1]),
        )
    ]
    resolutions = await workflow._resolve_batches(
        requests,
        options,
        progress_callback=None,
    )
    result_by_id = {result.source_id: result for result in existing}
    for resolution in resolutions:
        source = source_by_id.get(resolution.source_id)
        if source is None:
            continue
        accepted = workflow._accept_resolution(source, resolution)
        if accepted is not None:
            result_by_id[source.source_id] = accepted
    merged = [result_by_id[source.source_id] for source in sources]
    merged = assign_review_strata(sources, merged)
    merged = workflow._finalize(sources, merged, options)
    write_result_workbook(source_path, output_path, sources, merged)
    return sources, merged


def _parse_evidence(evidence: str) -> list[NameComponent]:
    components: list[NameComponent] = []
    cursor = 0
    for match in _EVIDENCE_COMPONENT.finditer(evidence):
        fragment = match.group("fragment").strip()
        components.append(
            NameComponent(
                source_fragment=fragment,
                full_name=match.group("full").strip(),
                korean_word=match.group("korean").strip(),
                origin=match.group("origin"),
                start=cursor,
                end=cursor + len(fragment),
            )
        )
        cursor += len(fragment)
    return components


def apply_review_corrections(
    *,
    source_path: str | Path,
    result_path: str | Path,
    mapping_path: str | Path,
    output_path: str | Path,
    corrections: list[dict[str, Any]],
) -> tuple[list, list[ColumnResult]]:
    sources = read_source_workbook(source_path)
    existing = read_existing_results(result_path)
    source_by_id = {source.source_id: source for source in sources}
    result_by_id = {result.source_id: result for result in existing}
    seen: set[str] = set()
    for correction in corrections:
        source_id = str(correction["source_id"])
        if source_id in seen:
            raise ValueError(f"중복 리뷰 보정 source_id: {source_id}")
        seen.add(source_id)
        source = source_by_id.get(source_id)
        if source is None:
            raise ValueError(f"원본에 없는 리뷰 보정 source_id: {source_id}")
        result_by_id[source_id] = _build_corrected_result(
            source.column_name,
            source_id,
            correction,
        )
    merged = [result_by_id[source.source_id] for source in sources]
    merged = assign_review_strata(sources, merged)
    glossary = MappingGlossary.from_xlsx(mapping_path)
    workflow = NamingWorkflow(glossary)
    options = WorkflowOptions(use_llm=False)
    merged = workflow._finalize(sources, merged, options)
    write_result_workbook(source_path, output_path, sources, merged)
    return sources, merged


def _build_corrected_result(
    column_name: str,
    source_id: str,
    correction: dict[str, Any],
) -> ColumnResult:
    full_name = str(correction["english_full_name"]).strip().upper()
    korean_name = re.sub(
        r"\s+",
        "",
        str(correction["korean_attribute_name"]).strip(),
    )
    if not full_name or not re.fullmatch(r"[가-힣0-9]+", korean_name):
        raise ValueError(f"잘못된 리뷰 보정 값: {source_id}")
    fragment = re.sub(r"[^A-Z0-9]", "", column_name.upper())
    component = NameComponent(
        source_fragment=fragment,
        full_name=full_name,
        korean_word=korean_name,
        origin="inference",
        start=0,
        end=len(fragment),
    )
    return ColumnResult(
        source_id=source_id,
        components=[component],
        english_full_name=full_name,
        korean_attribute_name=korean_name,
        status="검토필요",
        confidence=int(correction.get("confidence", 88)),
        evidence=f"{fragment}→{full_name}→{korean_name}[inference]",
        reason=str(correction.get("reason", "독립 리뷰 보정")),
        review_stratum="unmapped_inference",
    )
