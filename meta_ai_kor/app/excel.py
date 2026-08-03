from __future__ import annotations

import re
import shutil
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exceptions import InputWorkbookError
from app.glossary import MappingGlossary, normalize_header
from app.models import (
    ColumnResult,
    NameComponent,
    SourceRow,
)

SOURCE_HEADERS = [
    "스키마",
    "테이블명",
    "테이블설명",
    "컬럼순번",
    "컬럼명",
    "데이터타입",
    "데이터길이",
    "소수점",
    "PK여부",
    "NULL여부",
    "기본값",
    "컬럼설명",
]
RESULT_HEADERS = [
    "영문 Full Name",
    "한글속성명",
    "처리상태",
    "신뢰도",
    "변환근거",
]


def read_source_workbook(path: str | Path) -> list[SourceRow]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InputWorkbookError(f"입력 엑셀 파일을 열 수 없습니다: {exc}") from exc
    worksheet = workbook.active
    header_row, indexes = _find_source_header(worksheet)
    raw_headers = [
        "" if value is None else str(value)
        for value in next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )
    ]
    normalized_headers = [normalize_header(value) for value in raw_headers]
    if len(normalized_headers) != 12:
        workbook.close()
        raise InputWorkbookError(
            f"원본 컬럼은 12개여야 하지만 {len(normalized_headers)}개입니다."
        )
    rows: list[SourceRow] = []
    for excel_row, values in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_col=len(raw_headers),
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        column_name = _cell_text(values, indexes["컬럼명"])
        table_name = _cell_text(values, indexes["테이블명"])
        data_type = _cell_text(values, indexes["데이터타입"])
        if not column_name:
            workbook.close()
            raise InputWorkbookError(
                f"{excel_row}행의 필수 컬럼명 값이 비어 있습니다."
            )
        rows.append(
            SourceRow(
                source_id=f"row-{excel_row}",
                excel_row=excel_row,
                original_headers=raw_headers,
                original_values=list(values),
                schema_name=_optional_cell_text(values, indexes.get("스키마")),
                table_name=table_name,
                table_description=_cell_text(
                    values,
                    indexes["테이블설명"],
                ),
                column_ordinal=_cell_value(values, indexes["컬럼순번"]),
                column_name=column_name,
                data_type=data_type,
                column_description=_cell_text(
                    values,
                    indexes["컬럼설명"],
                ),
            )
        )
    workbook.close()
    if not rows:
        raise InputWorkbookError("처리할 컬럼 데이터가 없습니다.")
    return rows


def build_s1_baseline(
    sources: Iterable[SourceRow],
    glossary: MappingGlossary,
) -> list[ColumnResult]:
    results: list[ColumnResult] = []
    for source in sources:
        components: list[NameComponent] = []
        ambiguous = False
        unresolved = False
        cursor = 0
        for token in filter(None, source.column_name.split("_")):
            entry, entry_ambiguous = glossary.resolve(token, source)
            if entry is None:
                unresolved = True
                components.append(
                    NameComponent(
                        source_fragment=token,
                        full_name=token,
                        korean_word="미정",
                        origin="inference",
                        start=cursor,
                        end=cursor + len(token),
                    )
                )
            else:
                ambiguous = ambiguous or entry_ambiguous
                components.append(
                    NameComponent(
                        source_fragment=token,
                        full_name=entry.full_name,
                        korean_word=entry.korean_word,
                        origin="mapping",
                        start=cursor,
                        end=cursor + len(token),
                        occurrence_count=entry.occurrence_count,
                    )
                )
            cursor += len(token) + 1
        korean_words: list[str] = []
        for component in components:
            word = re.sub(r"[\s_]+", "", component.korean_word)
            if word and (not korean_words or korean_words[-1] != word):
                korean_words.append(word)
        confidence = 25 if unresolved else (65 if ambiguous else 80)
        results.append(
            ColumnResult(
                source_id=source.source_id,
                components=components,
                english_full_name=" ".join(
                    component.full_name for component in components
                ),
                korean_attribute_name="".join(korean_words) or "미정",
                status="검토필요",
                confidence=confidence,
                evidence=" | ".join(
                    (
                        f"{component.source_fragment}→"
                        f"{component.full_name}→{component.korean_word}"
                    )
                    for component in components
                ),
                reason="S1 사전 직접 일치 baseline",
                review_stratum=(
                    "unmapped_inference"
                    if unresolved
                    else (
                        "mapping_ambiguity"
                        if ambiguous
                        else "deterministic"
                    )
                ),
            )
        )
    return results


def write_result_workbook(
    input_path: str | Path,
    output_path: str | Path,
    sources: Iterable[SourceRow],
    results: Iterable[ColumnResult],
) -> None:
    source_list = list(sources)
    result_list = list(results)
    result_by_id = {result.source_id: result for result in result_list}
    if len(source_list) != len(result_list):
        raise InputWorkbookError(
            "입력 행 수와 결과 행 수가 일치하지 않습니다."
        )
    if set(result_by_id) != {source.source_id for source in source_list}:
        raise InputWorkbookError("입력과 결과의 source_id 집합이 다릅니다.")

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(input_path, destination)
    workbook = load_workbook(destination)
    worksheet = workbook.active
    worksheet.title = "한글속성명_결과"
    header_row, _ = _find_source_header(worksheet)
    start_column = 13
    for offset, header in enumerate(RESULT_HEADERS):
        target = worksheet.cell(row=header_row, column=start_column + offset)
        target.value = header
        _copy_header_style(worksheet.cell(row=header_row, column=12), target)
    for source in source_list:
        result = result_by_id[source.source_id]
        values = [
            result.english_full_name,
            result.korean_attribute_name,
            result.status,
            result.confidence,
            result.evidence,
        ]
        for offset, value in enumerate(values):
            cell = worksheet.cell(
                row=source.excel_row,
                column=start_column + offset,
                value=value,
            )
            cell.alignment = Alignment(
                horizontal="right" if offset == 3 else "left",
                vertical="center",
                wrap_text=offset == 4,
            )
        worksheet.cell(
            row=source.excel_row,
            column=start_column + 3,
        ).number_format = "0"

    widths = {
        "M": 40,
        "N": 30,
        "O": 14,
        "P": 10,
        "Q": 70,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    for table in worksheet.tables.values():
        table.ref = f"A{header_row}:Q{source_list[-1].excel_row}"

    review_results = [
        result
        for result in result_list
        if result.status in {"검토필요", "검증실패"}
    ]
    if review_results:
        _add_review_sheet(workbook, source_list, review_results)
    workbook.save(destination)


def _add_review_sheet(
    workbook,
    sources: list[SourceRow],
    results: list[ColumnResult],
) -> None:
    worksheet = workbook.create_sheet("검토필요")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    headers = [
        "원본행",
        "테이블명",
        "테이블설명",
        "컬럼명",
        "데이터타입",
        *RESULT_HEADERS,
        "검토사유",
        "추천확인포인트",
    ]
    worksheet.append(headers)
    source_by_id = {source.source_id: source for source in sources}
    for result in results:
        source = source_by_id[result.source_id]
        worksheet.append(
            [
                result.source_id,
                source.table_name,
                source.table_description,
                source.column_name,
                source.data_type,
                result.english_full_name,
                result.korean_attribute_name,
                result.status,
                result.confidence,
                result.evidence,
                ", ".join(result.validation_codes) or result.review_stratum,
                _review_hint(result),
            ]
        )
    fill = PatternFill("solid", fgColor="C65911")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [14, 28, 44, 28, 14, 40, 30, 14, 10, 70, 32, 48]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="right" if cell.column == 9 else "left",
                vertical="top",
                wrap_text=cell.column in {3, 10, 11, 12},
            )
    table = Table(
        displayName="KoreanColumnReviewNeeded",
        ref=f"A1:L{len(results) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _review_hint(result: ColumnResult) -> str:
    if result.review_stratum == "unmapped_inference":
        return "미해석 약어의 영문 원형과 한글 의미를 확인"
    if result.review_stratum == "mapping_ambiguity":
        return "다의 약어가 테이블 문맥에 맞는지 확인"
    if result.review_stratum == "segmentation_ambiguity":
        return "붙은 약어의 분해 경계와 의미 순서를 확인"
    return "Full Name과 한글속성명의 업무 의미 확인"


def _find_source_header(worksheet) -> tuple[int, dict[str, int]]:
    wanted = {"테이블명", "테이블설명", "컬럼순번", "컬럼명", "데이터타입", "컬럼설명"}
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(20, worksheet.max_row),
            values_only=True,
        ),
        start=1,
    ):
        indexes = {
            normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None
        }
        if wanted.issubset(indexes):
            return row_number, indexes
    raise InputWorkbookError(
        "입력 헤더에서 테이블명·테이블설명·컬럼순번·컬럼명·데이터타입·"
        "컬럼설명을 찾지 못했습니다."
    )


def _copy_header_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    else:
        target.fill = PatternFill("solid", fgColor="4472C4")
        target.font = Font(color="FFFFFF", bold=True)
        target.alignment = Alignment(horizontal="center", vertical="center")


def _cell_text(row: tuple[object, ...], index: int) -> str:
    value = _cell_value(row, index)
    return "" if value is None else str(value).strip()


def _optional_cell_text(
    row: tuple[object, ...],
    index: int | None,
) -> str | None:
    if index is None:
        return None
    value = _cell_text(row, index)
    return value or None


def _cell_value(row: tuple[object, ...], index: int):
    if index >= len(row):
        return None
    return row[index]


def _column_letter(index: int) -> str:
    value = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        value = chr(65 + remainder) + value
    return value

