from pathlib import Path

from openpyxl import load_workbook

from app.excel import (
    RESULT_HEADERS,
    build_s1_baseline,
    read_source_workbook,
    write_result_workbook,
)
from app.glossary import MappingGlossary


ROOT = Path(__file__).parents[2]
SOURCE_PATH = ROOT / "data" / "table_column_template_컬럼코멘트N.xlsx"
MAPPING_PATH = Path(__file__).parents[1] / "result.xlsx"


def test_real_source_and_mapping_match_benchmark():
    sources = read_source_workbook(SOURCE_PATH)
    glossary = MappingGlossary.from_xlsx(MAPPING_PATH)

    assert len(sources) == 694
    assert len(glossary) == 838
    assert len(sources[0].original_values) == 12


def test_write_result_preserves_original_values(tmp_path):
    sources = read_source_workbook(SOURCE_PATH)
    glossary = MappingGlossary.from_xlsx(MAPPING_PATH)
    results = build_s1_baseline(sources, glossary)
    output = tmp_path / "result.xlsx"

    write_result_workbook(SOURCE_PATH, output, sources, results)

    workbook = load_workbook(output, read_only=True, data_only=True)
    result_sheet = workbook["한글속성명_결과"]
    headers = [cell.value for cell in result_sheet[1]]
    assert headers[12:17] == RESULT_HEADERS
    assert result_sheet.max_row == 695
    for source in (sources[0], sources[200], sources[-1]):
        actual = [
            result_sheet.cell(source.excel_row, column).value
            for column in range(1, 13)
        ]
        assert [None if value == "" else value for value in actual] == [
            None if value == "" else value for value in source.original_values
        ]
    assert "검토필요" in workbook.sheetnames
    workbook.close()
