from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from app.excel import RESULT_HEADERS
from app.glossary import normalize_header

KOREAN_NAME_PATTERN = re.compile(r"^[가-힣0-9]+$")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--result", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    source_book = load_workbook(
        args.source,
        read_only=True,
        data_only=True,
    )
    result_book = load_workbook(
        args.result,
        read_only=True,
        data_only=True,
    )
    source_sheet = source_book.active
    result_sheet = result_book["한글속성명_결과"]
    changed_original_cells = 0
    source_rows = source_sheet.iter_rows(
        min_row=1,
        max_col=12,
        values_only=True,
    )
    result_source_rows = result_sheet.iter_rows(
        min_row=1,
        max_col=12,
        values_only=True,
    )
    for before_row, after_row in zip(
        source_rows,
        result_source_rows,
        strict=False,
    ):
        for before, after in zip(before_row, after_row, strict=True):
            if (None if before == "" else before) != (
                None if after == "" else after
            ):
                changed_original_cells += 1
    result_rows = result_sheet.iter_rows(
        min_row=1,
        max_col=17,
        values_only=True,
    )
    first_row = next(result_rows)
    headers = [normalize_header(value) for value in first_row[12:17]]
    expected_headers = [normalize_header(value) for value in RESULT_HEADERS]
    status = Counter()
    blank_full_name = 0
    blank_korean_name = 0
    invalid_korean_name = 0
    placeholder_name = 0
    blank_evidence = 0
    confidence_out_of_range = 0
    context_values: dict[tuple[object, ...], set[tuple[object, object]]] = (
        defaultdict(set)
    )
    for row in result_rows:
        full_name = str(row[12] or "").strip()
        korean_name = str(row[13] or "").strip()
        row_status = str(row[14] or "").strip()
        confidence = row[15]
        evidence = str(row[16] or "").strip()
        status[row_status] += 1
        blank_full_name += not bool(full_name)
        blank_korean_name += not bool(korean_name)
        invalid_korean_name += bool(korean_name) and not bool(
            KOREAN_NAME_PATTERN.fullmatch(korean_name)
        )
        placeholder_name += any(
            word in korean_name for word in ("미정", "불명", "알수없음")
        )
        blank_evidence += not bool(evidence)
        confidence_out_of_range += (
            not isinstance(confidence, (int, float))
            or not 0 <= confidence <= 100
        )
        context_key = (row[4], row[1], row[2], row[5])
        context_values[context_key].add((full_name, korean_name))
    inconsistent_contexts = sum(
        len(values) > 1 for values in context_values.values()
    )
    payload = {
        "source_row_count": source_sheet.max_row - 1,
        "result_row_count": result_sheet.max_row - 1,
        "row_count_matches": source_sheet.max_row == result_sheet.max_row,
        "changed_original_cell_count": changed_original_cells,
        "result_headers_match": headers == expected_headers,
        "blank_full_name_count": blank_full_name,
        "blank_korean_name_count": blank_korean_name,
        "invalid_korean_name_count": invalid_korean_name,
        "placeholder_name_count": placeholder_name,
        "blank_evidence_count": blank_evidence,
        "confidence_out_of_range_count": confidence_out_of_range,
        "inconsistent_context_count": inconsistent_contexts,
        "status_counts": dict(status),
    }
    source_book.close()
    result_book.close()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
