from __future__ import annotations

import re
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.excel import read_source_columns
from app.models import SourceColumn, WorkflowOptions
from app.rules import load_rules
from app.workflow import (
    KoreanCommentWorkflow,
    _deduplicate_sources,
    _deterministic_candidate,
    _recovery_events,
    _rules_to_synonym_groups,
    _semantic_terms,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
ACTUAL_INPUT = REPO_ROOT / "data" / "table_column_template_컬럼코멘트Y.xlsx"
INSURANCE_RULES_PATH = (
    Path(__file__).resolve().parents[1]
    / "config"
    / "rules"
    / "examples"
    / "insurance.yaml"
)


class UnavailableModel:
    """Exercise the production deterministic recovery path without test data."""

    async def generate(self, sources, risks=None):
        raise RuntimeError("실제 로컬 LLM 엔드포인트 연결 실패 모사")

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        return []


def _insurance_rules():
    return load_rules(INSURANCE_RULES_PATH)


def test_glossary_translation_and_unknown_token_evidence() -> None:
    rules = _insurance_rules()
    glossary = rules.glossary_lookup()
    semantic_terms = _semantic_terms(rules)

    tpms = _deterministic_candidate(
        SourceColumn(
            source_id="tpms",
            column_name="TPMS_YN",
            column_description="TPMS여부",
        ),
        glossary=glossary,
        semantic_terms=semantic_terms,
    )
    assert tpms.korean_attribute_name == "타이어공기압감지장치여부"
    assert tpms.action.value == "rewrite"

    unknown = _deterministic_candidate(
        SourceColumn(
            source_id="unknown",
            column_name="XYZ_YN",
            column_description="XYZ여부",
        ),
        glossary=glossary,
        semantic_terms=semantic_terms,
    )
    assert unknown.korean_attribute_name == "XYZ여부"
    assert unknown.action.value == "rewrite"
    assert unknown.confidence == 55
    assert "확정하지 못한 영문" in unknown.review_reasons[0]


def test_deterministic_candidate_is_domain_neutral_without_rules() -> None:
    # With no glossary, TPMS is now an unknown token — the deterministic path
    # must surface it as review evidence instead of secretly translating it.
    tpms = _deterministic_candidate(
        SourceColumn(
            source_id="tpms",
            column_name="TPMS_YN",
            column_description="TPMS여부",
        )
    )
    assert tpms.korean_attribute_name == "TPMS여부"
    assert tpms.action.value == "rewrite"
    assert tpms.confidence == 55
    assert any("확정하지 못한 영문" in reason for reason in tpms.review_reasons)


def test_ordinal_prefix_is_left_for_the_llm_to_interpret() -> None:
    # The deterministic path no longer contains any column-name-keyed ordinal
    # rewrite. Column descriptions with ``제N`` stay verbatim so the LLM (or
    # a downstream reviewer) decides the correct Korean surface form.
    result = _deterministic_candidate(
        SourceColumn(
            source_id="row-250",
            column_name="SCD_INS_TRM_APPRM",
            column_description="제2보험기간적용보험료",
        )
    )
    assert result.korean_attribute_name == "제2보험기간적용보험료"
    assert result.action.value == "keep"


def test_review_recovery_events_are_sorted_by_round_before_code() -> None:
    aliases = {
        "representative": [
            SourceColumn(
                source_id="row-2",
                column_name="COL",
                column_description="테스트",
            )
        ]
    }

    events = _recovery_events(
        aliases,
        {},
        [
            (2, "representative", "timeout"),
            (1, "representative", "unexpected_error"),
        ],
    )

    assert [event["round"] for event in events] == ["1", "2"]


def test_actual_risky_duplicate_is_partitioned_by_table_context() -> None:
    sources = read_source_columns(ACTUAL_INPUT)

    representatives, aliases = _deduplicate_sources(sources)

    # The source has 921 basic column-name/description pairs.  Only the real
    # OLD_CATCD pair occurs in two distinct table contexts and therefore needs
    # two independently reviewable representatives.
    assert len(representatives) == 922
    old_representatives = [
        source
        for source in representatives
        if source.column_name == "OLD_CATCD"
        and source.column_description == "OLD차종코드"
    ]
    assert [source.source_id for source in old_representatives] == [
        "row-865",
        "row-923",
    ]
    assert all(len(aliases[source.source_id]) == 1 for source in old_representatives)


@pytest.mark.asyncio
async def test_actual_1195_row_workbook_end_to_end_with_insurance_preset() -> None:
    output_dir = Path(__file__).resolve().parents[1] / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"test-actual-{uuid4().hex}.xlsx"
    workflow = KoreanCommentWorkflow(UnavailableModel(), rules=_insurance_rules())

    try:
        summary = await workflow.run(
            ACTUAL_INPUT,
            output,
            WorkflowOptions(max_review_rounds=0),
        )

        assert summary.source_count == 1195
        assert summary.recovery_stats["generation_fallback_count"] > 0
        assert summary.recovery_stats["generation_unexpected_error_count"] > 0
        assert summary.recovery_stats["review_failure_count"] == 0
        assert output.exists()

        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            assert workbook.sheetnames == ["한글속성명_결과", "검토필요"]
            sheet = workbook["한글속성명_결과"]
            headers = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]
            assert headers[-6:] == [
                "한글속성명",
                "처리상태",
                "신뢰도",
                "처리방식",
                "변환근거",
                "검토사유",
            ]
            assert sheet.max_row == 1196
            assert sheet.max_column == 18

            name_index = headers.index("컬럼명 (*)")
            description_index = headers.index("컬럼설명")
            korean_index = headers.index("한글속성명")
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            by_pair = {
                (row[name_index], row[description_index]): row[korean_index]
                for row in rows
            }

            # Glossary-driven translations from the insurance preset survive the
            # deterministic fallback path with the LLM disabled.
            expected = {
                ("FY_YR", "FY년도"): "회계년도",
                ("SMS_RCV_YN", "SMS수신여부"): "문자메시지수신여부",
                ("RH_TYCD", "RH형태코드"): "리서스인자형태코드",
                ("SOFA_CR_YN", "SOFA차량여부"): "주한미군지위협정적용차량여부",
                ("TMR_MNTH_CNV_HMS", "TMR월통화시간"): "텔레마케터월통화시간",
                ("SP_YN", "SP여부"): "설계사여부",
                # Slash-only cases: the first alternative is picked as a
                # placeholder and the row is flagged for review.
                ("CHSNO_OR_TMPNO", "차대번호/임시번호"): "차대번호",
                # Synonym-group decisions land on the frequency-winning surface.
                ("PYM_CT", "납부횟수"): "납입횟수",
                ("USDCR_RT", "중고차율"): "중고차요율",
                ("CR_TYCD", "차형태코드"): "차량형태코드",
                ("LMIT_TRT_RT", "한정특약율"): "한정특약율",
                ("AGE_TRT_RT", "연령특약요율"): "연령특약율",
                # Generic ``등급등급→등급`` duplicate-suffix simplification.
                ("ONFML_PF_GRD_GRDCD", "한가족우대등급등급코드"): "한가족우대등급코드",
            }
            for pair, target in expected.items():
                assert by_pair[pair] == target, pair

            for row in rows:
                source = str(row[description_index])
                result = str(row[korean_index])
                assert result
                assert not re.search(r"[A-Za-z]", result.replace("ID", ""))
                assert not re.search(r"[^가-힣ㄱ-ㅎㅏ-ㅣ0-9A-Za-z]", result)
                assert re.findall(r"[0-9]+", source) == re.findall(r"[0-9]+", result)
        finally:
            workbook.close()
    finally:
        output.unlink(missing_ok=True)


def test_rules_to_synonym_groups_and_semantic_terms_derive_from_rules() -> None:
    rules = _insurance_rules()

    groups = _rules_to_synonym_groups(rules)
    terms = _semantic_terms(rules)

    assert {group.group_id for group in groups} == {
        "payment-action",
        "used-car-rate",
        "vehicle-form",
        "special-contract-rate",
    }
    # Longer surfaces come first so the segmentation prefers the maximal match.
    assert list(terms) == sorted(terms, key=lambda value: (-len(value), value))
    assert "중고차요율" in terms
