from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models import GenerationResult, ProcessingAction, SourceColumn, WorkflowOptions
from app.validation import derive_processing_status
from app.workflow import KoreanCommentWorkflow, _mark_review_failure, _recovery_stats


REPO_ROOT = Path(__file__).resolve().parents[2]
ACTUAL_INPUT = REPO_ROOT / "data" / "table_column_template_컬럼코멘트Y.xlsx"


class LowConfidenceThenReviewedModel:
    def __init__(self) -> None:
        self.review_calls = 0
        self.review_batch_sizes = []
        self.risk_batches = []
        self.terminology_contexts = []

    async def generate(self, sources, risks=None):
        self.risk_batches.append((list(sources), list(risks or [])))
        return [
            GenerationResult(
                source_id=source.source_id,
                original_description=source.column_description,
                korean_attribute_name="주한미군차량여부",
                action=ProcessingAction.REWRITE,
                confidence=70,
                reason="SOFA를 자동차정보 문맥에서 주한미군으로 한글화",
                semantic_units=["주한미군차량여부"],
            )
            for source in sources
            if source.column_name == "SOFA_CR_YN"
        ]

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        self.review_calls += 1
        self.review_batch_sizes.append(len(sources))
        self.terminology_contexts.append(list(terminology_context or []))
        return [
            result.model_copy(
                update={
                    "confidence": 99,
                    "reason": result.reason + "; 독립 재검토에서 문맥 확인",
                }
            )
            for result in current_results
        ]


@pytest.mark.asyncio
async def test_actual_low_confidence_row_is_reviewed_without_regressing_others() -> None:
    output_dir = Path(__file__).resolve().parents[1] / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"test-review-{uuid4().hex}.xlsx"
    model = LowConfidenceThenReviewedModel()
    try:
        result = await KoreanCommentWorkflow(model).run(
            ACTUAL_INPUT,
            output,
            WorkflowOptions(max_review_rounds=2),
        )
        assert model.review_calls >= 2
        assert max(model.review_batch_sizes) <= 25
        assert result.review_rounds == 2
        assert result.source_count == 1195
        assert result.validation_report.is_valid
        assert result.validation_failed_count == 0
        assert result.recovery_stats["generation_fallback_count"] > 0
        assert result.recovery_stats["generation_rejected_result_count"] > 0
        assert result.recovery_stats["review_failure_count"] == 0
        assert all(
            [source.source_id for source in sources]
            == [risk.source_id for risk in risks]
            for sources, risks in model.risk_batches
        )
        assert any(model.terminology_contexts)
        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            sheet = workbook["한글속성명_결과"]
            assert sheet["M190"].value == "주한미군지위협정적용차량여부"
            assert sheet["N28"].value == "자동확정"
            assert sheet["N190"].value == "검토필요"
        finally:
            workbook.close()
    finally:
        output.unlink(missing_ok=True)


class FailingReviewModel:
    async def generate(self, sources, risks=None):
        return []

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        raise RuntimeError("리뷰 엔드포인트 장애")


class CapturingBatchedReviewModel:
    def __init__(self) -> None:
        self.calls = []

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        self.calls.append(
            {
                "source_ids": [source.source_id for source in sources],
                "issues": list(issues),
            }
        )
        return list(current_results)


@pytest.mark.asyncio
async def test_review_uses_configured_batches_and_threshold() -> None:
    sources = [
        SourceColumn(
            source_id=f"row-{index}",
            column_name=f"COL_{index}",
            column_description=f"테스트{index}",
        )
        for index in range(1, 8)
    ]
    results = [
        GenerationResult(
            source_id=source.source_id,
            original_description=source.column_description,
            korean_attribute_name=source.column_description,
            action=ProcessingAction.KEEP,
            confidence=92,
            reason="정책을 충족하는 컬럼설명을 원문 그대로 유지",
            semantic_units=[source.column_description],
        )
        for source in sources
    ]
    model = CapturingBatchedReviewModel()
    reviewed, failure = await KoreanCommentWorkflow(model)._review(
        sources,
        results,
        [],
        [],
        {source.source_id for source in sources},
        1,
        WorkflowOptions(
            batch_size=2,
            max_concurrency=2,
            auto_confirm_threshold=95,
        ),
    )

    assert failure == {}
    assert reviewed is not None and len(reviewed) == 7
    assert sorted(len(call["source_ids"]) for call in model.calls) == [1, 2, 2, 2]
    assert all(
        any(
            issue.code == "low_confidence"
            and issue.details["auto_confirm_threshold"] == 95
            for issue in call["issues"]
        )
        for call in model.calls
    )


class PartiallyFailingBatchedReviewModel:
    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        if any(source.source_id == "row-3" for source in sources):
            raise RuntimeError("선택 배치 장애")
        return [
            result.model_copy(update={"confidence": 99})
            for result in current_results
        ]


@pytest.mark.asyncio
async def test_review_preserves_successful_batches_when_one_batch_fails() -> None:
    sources = [
        SourceColumn(
            source_id=f"row-{index}",
            column_name=f"COL_{index}",
            column_description=f"테스트{index}",
        )
        for index in range(1, 5)
    ]
    results = [
        GenerationResult(
            source_id=source.source_id,
            original_description=source.column_description,
            korean_attribute_name=source.column_description,
            action=ProcessingAction.KEEP,
            confidence=80,
            reason="원문 유지",
            semantic_units=[source.column_description],
        )
        for source in sources
    ]

    reviewed, failures = await KoreanCommentWorkflow(
        PartiallyFailingBatchedReviewModel()
    )._review(
        sources,
        results,
        [],
        [],
        {source.source_id for source in sources},
        1,
        WorkflowOptions(batch_size=2, max_concurrency=2),
    )

    by_id = {result.source_id: result for result in reviewed}
    assert failures == {
        "row-3": "unexpected_error",
        "row-4": "unexpected_error",
    }
    assert by_id["row-1"].confidence == 99
    assert by_id["row-2"].confidence == 99
    assert by_id["row-3"].confidence == 80
    assert by_id["row-4"].confidence == 80


class TransientFirstReviewBatchFailureModel:
    def __init__(self) -> None:
        self.failed_once = False

    async def generate(self, sources, risks=None):
        return []

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        if not self.failed_once:
            self.failed_once = True
            raise RuntimeError("일시적 첫 배치 장애")
        return [
            result.model_copy(update={"confidence": 99})
            for result in current_results
        ]


@pytest.mark.asyncio
async def test_transient_review_failure_remains_in_recovery_history() -> None:
    output_dir = Path(__file__).resolve().parents[1] / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"test-review-history-{uuid4().hex}.xlsx"
    try:
        result = await KoreanCommentWorkflow(
            TransientFirstReviewBatchFailureModel()
        ).run(
            ACTUAL_INPUT,
            output,
            WorkflowOptions(
                batch_size=5,
                max_concurrency=3,
                max_review_rounds=2,
            ),
        )

        assert result.validation_report.is_valid
        assert result.review_rounds == 2
        assert result.recovery_stats["review_failure_count"] > 0
        assert result.recovery_stats["review_failure_source_count"] > 0
        assert result.recovery_stats["review_unresolved_failure_count"] == 0
        assert any(
            event["stage"] == "review"
            and event["round"] == "1"
            and event["code"] == "unexpected_error"
            for event in result.recovery_events
        )
    finally:
        output.unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_actual_review_failure_is_traced_without_losing_output() -> None:
    output_dir = Path(__file__).resolve().parents[1] / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"test-review-failure-{uuid4().hex}.xlsx"
    try:
        result = await KoreanCommentWorkflow(FailingReviewModel()).run(
            ACTUAL_INPUT,
            output,
            WorkflowOptions(max_review_rounds=2),
        )
        assert result.review_rounds == 2
        assert result.validation_report.is_valid
        assert result.validation_failed_count > 0
        assert result.recovery_stats["generation_fallback_count"] > 0
        assert result.recovery_stats["review_failure_count"] > 0
        assert result.recovery_stats["review_unexpected_error_count"] > 0
        assert any(
            event["stage"] == "review"
            and event["code"] == "unexpected_error"
            for event in result.recovery_events
        )

        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            review_sheet = workbook["검토필요"]
            reasons = [
                str(row[0] or "")
                for row in review_sheet.iter_rows(
                    min_col=13,
                    max_col=13,
                    values_only=True,
                )
            ]
            assert any("로컬 LLM 리뷰 " in reason for reason in reasons)
        finally:
            workbook.close()
    finally:
        output.unlink(missing_ok=True)


def test_policy_rejection_is_tracked_separately_from_execution_failure() -> None:
    stats = _recovery_stats(
        {"row-1": 2, "row-2": 1},
        {},
        [
            (1, "row-1", "rejected_result"),
            (2, "row-1", "rejected_result"),
            (1, "row-2", "timeout"),
        ],
        {"row-1": "rejected_result", "row-2": "timeout"},
    )

    assert stats["review_failure_count"] == 1
    assert stats["review_failure_source_count"] == 1
    assert stats["review_unresolved_failure_count"] == 1
    assert stats["review_rejected_result_count"] == 4
    assert stats["review_rejected_result_source_count"] == 2
    assert stats["review_unresolved_rejected_result_count"] == 2

    empty = _recovery_stats({"row-1": 1}, {}, [], {})
    assert empty["review_rejected_result_count"] == 0


class AlternateValidReviewModel:
    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        return [
            result.model_copy(
                update={
                    "korean_attribute_name": "시험",
                    "action": ProcessingAction.REWRITE,
                    "confidence": 99,
                    "reason": "대체 유효 후보",
                    "semantic_units": ["시험"],
                }
            )
            for result in current_results
        ]


@pytest.mark.asyncio
async def test_successful_policy_rejection_clears_prior_execution_failure() -> None:
    source = SourceColumn(
        source_id="row-1",
        column_name="TEST",
        column_description="테스트",
    )
    current = GenerationResult(
        source_id=source.source_id,
        original_description=source.column_description,
        korean_attribute_name=source.column_description,
        action=ProcessingAction.KEEP,
        confidence=85,
        reason="원문 유지",
        semantic_units=[source.column_description],
        review_reasons=[
            "로컬 LLM 리뷰 시간 초과로 현재 검증 결과를 유지"
        ],
    )

    reviewed, failures = await KoreanCommentWorkflow(
        AlternateValidReviewModel()
    )._review(
        [source],
        [current],
        [],
        [],
        {source.source_id},
        2,
        WorkflowOptions(),
    )
    marked = _mark_review_failure(
        reviewed,
        {source.source_id},
        "rejected_result",
    )

    assert failures == {source.source_id: "rejected_result"}
    assert all("시간 초과" not in reason for reason in marked[0].review_reasons)
    assert derive_processing_status(marked[0], []).value == "검토필요"


class RejectedFirstRoundThenAcceptedModel:
    def __init__(self) -> None:
        self.rounds: list[int] = []

    async def generate(self, sources, risks=None):
        return []

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        self.rounds.append(review_round)
        if review_round == 1:
            return [
                result.model_copy(update={"original_description": "잘못된원문"})
                for result in current_results
            ]
        return [result.model_copy(update={"confidence": 99}) for result in current_results]


@pytest.mark.asyncio
async def test_all_policy_rejections_use_remaining_review_round() -> None:
    output_dir = Path(__file__).resolve().parents[1] / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"test-review-retry-{uuid4().hex}.xlsx"
    model = RejectedFirstRoundThenAcceptedModel()
    try:
        result = await KoreanCommentWorkflow(model).run(
            ACTUAL_INPUT,
            output,
            WorkflowOptions(max_review_rounds=2),
        )

        assert result.review_rounds == 2
        assert set(model.rounds) == {1, 2}
        assert result.validation_failed_count == 0
        assert result.recovery_stats["review_failure_count"] == 0
        assert result.recovery_stats["review_rejected_result_count"] > 0
        assert result.recovery_stats["review_unresolved_rejected_result_count"] == 0
    finally:
        output.unlink(missing_ok=True)
