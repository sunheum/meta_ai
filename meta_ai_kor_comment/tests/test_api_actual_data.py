from __future__ import annotations

import io
import time
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import Settings
from app.main import XLSX_MEDIA_TYPE, create_app
from app.workflow import KoreanCommentWorkflow


REPO_ROOT = Path(__file__).resolve().parents[2]
ACTUAL_INPUT = REPO_ROOT / "data" / "table_column_template_컬럼코멘트Y.xlsx"
RESULTS_DIR = Path(__file__).resolve().parents[1] / "results"


class DeterministicModel:
    async def generate(self, sources, risks=None):
        return []

    async def review(
        self,
        sources,
        current_results,
        issues,
        review_round,
        terminology_context=None,
    ):
        return []


def _client() -> TestClient:
    settings = Settings(
        results_dir=str(RESULTS_DIR),
        progress_heartbeat_seconds=0.01,
        default_max_review_rounds=0,
    )
    return TestClient(
        create_app(
            settings=settings,
            workflow=KoreanCommentWorkflow(DeterministicModel()),
        )
    )


def _fault_injected_actual_workbook() -> bytes:
    """Copy the real workbook and inject one unresolved translation fault."""

    workbook = load_workbook(ACTUAL_INPUT)
    try:
        worksheet = workbook["테이블_컬럼_정보"]
        header_row = next(
            row_index
            for row_index in range(1, min(worksheet.max_row, 20) + 1)
            if any(
                str(worksheet.cell(row_index, column_index).value).strip()
                in {"컬럼설명", "컬럼 설명"}
                for column_index in range(1, worksheet.max_column + 1)
                if worksheet.cell(row_index, column_index).value is not None
            )
        )
        description_column = next(
            column_index
            for column_index in range(1, worksheet.max_column + 1)
            if str(worksheet.cell(header_row, column_index).value).strip()
            in {"컬럼설명", "컬럼 설명"}
        )
        worksheet.cell(header_row + 1, description_column).value = "XYZ여부"
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    finally:
        workbook.close()


def test_health_and_sync_endpoint_with_actual_workbook() -> None:
    with _client() as client:
        health = client.get("/health")
        assert health.status_code == 200
        assert health.json()["status"] == "ok"

        with ACTUAL_INPUT.open("rb") as stream:
            response = client.post(
                "/v1/comment-korean-column-names",
                files={
                    "file": (
                        ACTUAL_INPUT.name,
                        stream,
                        XLSX_MEDIA_TYPE,
                    )
                },
                data={"max_review_rounds": "0"},
            )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
        assert response.headers["x-source-count"] == "1195"
        assert response.headers["x-result-count"] == "1195"
        assert response.headers["x-validation-errors"] == "0"
        assert int(response.headers["x-generation-fallback-count"]) > 0
        assert response.headers["x-review-failure-count"] == "0"

        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        try:
            assert workbook.sheetnames == ["한글속성명_결과", "검토필요"]
            assert workbook["한글속성명_결과"].max_row == 1196
        finally:
            workbook.close()


def test_async_job_status_and_progress_with_actual_workbook() -> None:
    result_path: Path | None = None
    try:
        with _client() as client:
            with ACTUAL_INPUT.open("rb") as stream:
                created = client.post(
                    "/v1/comment-korean-column-names/jobs",
                    files={
                        "file": (
                            ACTUAL_INPUT.name,
                            stream,
                            XLSX_MEDIA_TYPE,
                        )
                    },
                    data={"max_review_rounds": "0"},
                )
            assert created.status_code == 202
            job_id = created.json()["job_id"]

            snapshot = None
            for _ in range(200):
                response = client.get(
                    f"/v1/comment-korean-column-names/jobs/{job_id}"
                )
                assert response.status_code == 200
                snapshot = response.json()
                if snapshot["status"] in {"completed", "failed"}:
                    break
                time.sleep(0.025)

            assert snapshot is not None
            assert snapshot["status"] == "completed"
            metadata = snapshot["result_metadata"]
            assert metadata["source_count"] == 1195
            assert metadata["result_count"] == 1195
            assert metadata["validation_stats"]["error_count"] == 0
            assert metadata["recovery_stats"]["generation_fallback_count"] > 0
            assert metadata["recovery_stats"]["generation_missing_result_count"] > 0
            assert metadata["recovery_stats"]["review_failure_count"] == 0
            assert metadata["recovery_events"]
            result_path = Path(metadata["result_path"])
            assert result_path.exists()

            progress = client.get(
                f"/v1/comment-korean-column-names/jobs/{job_id}/progress"
            )
            assert progress.status_code == 200
            assert "완료" in progress.text
            assert "XLSX 저장 경로" in progress.text
    finally:
        if result_path is not None:
            result_path.unlink(missing_ok=True)


def test_unresolved_actual_fault_is_preserved_as_partial_sync_and_async_result() -> None:
    content = _fault_injected_actual_workbook()
    result_path: Path | None = None
    try:
        with _client() as client:
            sync_response = client.post(
                "/v1/comment-korean-column-names",
                files={"file": (ACTUAL_INPUT.name, content, XLSX_MEDIA_TYPE)},
                data={"max_review_rounds": "0"},
            )

            assert sync_response.status_code == 200
            assert sync_response.headers["x-source-count"] == "1195"
            assert sync_response.headers["x-result-count"] == "1195"
            assert sync_response.headers["x-partial-result"] == "true"
            assert sync_response.headers["x-validation-errors"] == "1"

            workbook = load_workbook(
                io.BytesIO(sync_response.content), read_only=True, data_only=True
            )
            try:
                result_sheet = workbook["한글속성명_결과"]
                headers = [
                    cell.value
                    for cell in next(
                        result_sheet.iter_rows(min_row=1, max_row=1)
                    )
                ]
                status_index = headers.index("처리상태")
                statuses = [
                    row[status_index]
                    for row in result_sheet.iter_rows(min_row=2, values_only=True)
                ]
                assert statuses.count("검증실패") == 1
                assert workbook["검토필요"].max_row >= 2
            finally:
                workbook.close()

            created = client.post(
                "/v1/comment-korean-column-names/jobs",
                files={"file": (ACTUAL_INPUT.name, content, XLSX_MEDIA_TYPE)},
                data={"max_review_rounds": "0"},
            )
            assert created.status_code == 202
            job_id = created.json()["job_id"]

            snapshot = None
            for _ in range(200):
                status_response = client.get(
                    f"/v1/comment-korean-column-names/jobs/{job_id}"
                )
                assert status_response.status_code == 200
                snapshot = status_response.json()
                if snapshot["status"] in {"completed", "failed"}:
                    break
                time.sleep(0.025)

            assert snapshot is not None
            assert snapshot["status"] == "completed"
            metadata = snapshot["result_metadata"]
            assert metadata["source_count"] == 1195
            assert metadata["result_count"] == 1195
            assert metadata["is_partial"] is True
            assert metadata["validation_failed_count"] == 1
            assert metadata["validation_stats"]["error_count"] == 1
            result_path = Path(metadata["result_path"])
            assert result_path.exists()
    finally:
        if result_path is not None:
            result_path.unlink(missing_ok=True)
