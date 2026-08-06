from pathlib import Path

from openpyxl import load_workbook

from app.models import RiskLevel, SourceColumn
from app.normalization import (
    can_keep_description,
    classify_description,
    digit_sequences,
    digits_are_preserved,
    forbidden_characters,
    invalid_english_tokens,
    source_dedup_key,
)


def test_clean_korean_and_exact_id_can_be_kept() -> None:
    assert can_keep_description("고객ID")
    assert invalid_english_tokens("고객ID") == ()
    assert forbidden_characters("제2고객ID") == ()

    risk = classify_description("제2고객ID")

    assert risk.level is RiskLevel.MEDIUM
    assert risk.codes == ["numeric_sensitive"]
    assert not risk.requires_generation
    assert not risk.requires_review


def test_all_english_except_exact_uppercase_id_requires_translation() -> None:
    for value in ("FY년도", "SMS수신여부", "고객id", "TPMS여부"):
        risk = classify_description(value)

        assert "english_translation_required" in risk.codes
        assert risk.requires_generation
        assert risk.level is RiskLevel.HIGH


def test_digit_sequences_must_match_exact_value_and_order() -> None:
    assert digit_sequences("제2약관23조") == ("2", "23")
    assert digits_are_preserved("제2약관23조", "제2약관23조여부")
    assert not digits_are_preserved("제2약관23조", "제23약관2조")
    assert not digits_are_preserved("518코드", "오일팔코드")


def test_slash_and_symbols_are_explicit_high_risks() -> None:
    risk = classify_description("차대번호/임시번호")

    assert risk.codes == ["slash_ambiguity", "special_symbol"]
    assert risk.symbols == ["/"]
    assert risk.level is RiskLevel.HIGH
    assert risk.requires_generation
    assert risk.requires_review
    assert not can_keep_description("차대번호/임시번호")


def test_non_korean_non_english_character_is_rejected() -> None:
    risk = classify_description("顧客번호")

    assert "unsupported_character" in risk.codes
    assert forbidden_characters("顧客번호") == ("顧", "客")


def test_dedup_key_preserves_description_semantics() -> None:
    first = SourceColumn(
        source_id="row-2",
        column_name=" use_yn ",
        column_description="사용 여부",
    )
    second = SourceColumn(
        source_id="row-3",
        column_name="USE_YN",
        column_description="사용   여부",
    )

    assert source_dedup_key(first) == source_dedup_key(second)
    assert source_dedup_key(first) == ("USE_YN", "사용 여부")


def test_actual_input_policy_profile_matches_handoff_baseline() -> None:
    input_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "table_column_template_컬럼코멘트Y.xlsx"
    )
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = workbook["테이블_컬럼_정보"]
        rows = sheet.iter_rows(values_only=True)
        headers = [
            str(value).replace("(*)", "").strip() if value is not None else ""
            for value in next(rows)
        ]
        column_name_index = headers.index("컬럼명")
        description_index = headers.index("컬럼설명")
        sources = [
            SourceColumn(
                source_id=f"row-{excel_row}",
                column_name=str(row[column_name_index] or ""),
                column_description=str(row[description_index] or ""),
            )
            for excel_row, row in enumerate(rows, start=2)
        ]
    finally:
        workbook.close()

    assessments = [
        classify_description(source.column_description) for source in sources
    ]
    assert len(sources) == 1_195
    assert all(source.column_description for source in sources)
    assert sum(bool(assessment.english_tokens) for assessment in assessments) == 51
    assert sum(
        "english_translation_required" in assessment.codes
        for assessment in assessments
    ) == 28
    assert sum("numeric_sensitive" in assessment.codes for assessment in assessments) == 45
    assert sum("slash_ambiguity" in assessment.codes for assessment in assessments) == 4
    assert len({source_dedup_key(source) for source in sources}) == 921
