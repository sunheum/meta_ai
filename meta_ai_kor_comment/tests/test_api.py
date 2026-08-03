import asyncio
from io import BytesIO
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app.config import Settings
from app.excel import read_source_columns, write_result_workbook
from app.main import create_app
from app.models import (
    KoreanAttributeResult,
    ProcessingAction,
    ProcessingStatus,
    ProgressEvent,
    ValidationReport,
    WorkflowResult,
)


ROOT = Path(__file__).resolve().parents[2]
ACTUAL_INPUT = ROOT / "data" / "table_column_template_컬럼코멘트Y.xlsx"


class ActualDataWorkflow:
    def __init__(self, delay_seconds: float = 0) -> None:
        self.delay_seconds = delay_seconds
        self.options = None

    async def run(
        self,
        input_path,
        output_path,
        options,
        progress_callback=None,
    ) -> WorkflowResult:
        self.options = options
        sources = read_source_columns(input_path)
        if progress_callback:
            await progress_callback(
                ProgressEvent(
                    stage="input",
                    stage_percent=100,
                    overall_percent=10,
                    message=f"실제 입력 {len(sources):,}행을 읽었습니다.",
                )
            )
            await progress_callback(
                ProgressEvent(
                    stage="normalize",
                    stage_percent=100,
                    overall_percent=25,
                    message="정규화 완료",
                )
            )
            await progress_callback(
                ProgressEvent(
                    stage="generate",
                    stage_percent=50,
                    overall_percent=45,
                    message="생성 중",
                )
            )
        if self.delay_seconds:
            await asyncio.sleep(self.delay_seconds)

        results = [
            KoreanAttributeResult(
                source_id=source.source_id,
                original_description=source.column_description,
                korean_attribute_name=source.column_description,
                action=ProcessingAction.KEEP,
                confidence=100,
                status=ProcessingStatus.AUTO_CONFIRMED,
            )
            for source in sources
        ]
        if progress_callback:
            await progress_callback(
                ProgressEvent(
                    stage="validate",
                    stage_percent=100,
                    overall_percent=90,
                    message="결정적 검증 통과",
                )
            )
        write_result_workbook(input_path, output_path, sources, results)
        if progress_callback:
            await progress_callback(
                ProgressEvent(
                    stage="output",
                    stage_percent=100,
                    overall_percent=100,
                    message="XLSX 생성 완료",
                )
            )
        return WorkflowResult(
            output_path=str(output_path),
            source_count=len(sources),
            auto_confirmed_count=len(sources),
            validation_report=ValidationReport(
                is_valid=True,
                issues=[],
                stats={"error_count": 0, "warning_count": 0},
            ),
        )


def test_single_request_uses_actual_workbook_and_returns_preserved_xlsx() -> None:
    workflow = ActualDataWorkflow()
    app = create_app(settings=Settings(), workflow=workflow)

    async def request() -> httpx.Response:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test"
        ) as client:
            return await client.post(
                "/v1/comment-korean-column-names",
                files={
                    "file": (
                        ACTUAL_INPUT.name,
                        ACTUAL_INPUT.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"auto_confirm_threshold": "95"},
            )

    response = asyncio.run(request())

    assert response.status_code == 200
    assert response.headers["x-source-count"] == "1195"
    assert response.headers["x-result-count"] == "1195"
    assert response.headers["x-partial-result"] == "false"
    assert response.headers["x-generation-fallback-count"] == "0"
    assert response.headers["x-review-failure-count"] == "0"
    assert workflow.options.auto_confirm_threshold == 95
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        assert workbook.sheetnames == ["한글속성명_결과", "검토필요"]
        result_sheet = workbook["한글속성명_결과"]
        assert result_sheet.max_row == 1196
        assert result_sheet.max_column == 18
    finally:
        workbook.close()


def test_job_api_streams_heartbeat_and_saves_actual_result() -> None:
    results_dir = Path(__file__).resolve().parents[1] / "results"
    app = create_app(
        settings=Settings(
            progress_heartbeat_seconds=0.01,
            results_dir=str(results_dir),
        ),
        workflow=ActualDataWorkflow(delay_seconds=0.05),
    )

    async def request():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test", timeout=30
        ) as client:
            created = await client.post(
                "/v1/comment-korean-column-names/jobs",
                files={
                    "file": (
                        ACTUAL_INPUT.name,
                        ACTUAL_INPUT.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            job_id = created.json()["job_id"]
            progress = await client.get(
                f"/v1/comment-korean-column-names/jobs/{job_id}/progress"
            )
            status = await client.get(
                f"/v1/comment-korean-column-names/jobs/{job_id}"
            )
            return created, progress, status

    saved_path = None
    try:
        created, progress, status = asyncio.run(request())

        assert created.status_code == 202
        assert "result_url" not in created.json()
        assert progress.status_code == 200
        assert "정규화" in progress.text
        assert "LLM 생성" in progress.text
        assert "결정적 검증 통과" in progress.text
        assert "연결 유지" in progress.text
        assert "단계별 소요시간" in progress.text
        assert status.status_code == 200
        metadata = status.json()["result_metadata"]
        assert metadata["source_count"] == 1195
        assert metadata["result_count"] == 1195
        assert metadata["recovery_stats"] == {}
        assert metadata["recovery_events"] == []
        saved_path = Path(metadata["result_path"])
        assert saved_path == results_dir / f"{created.json()['job_id']}.xlsx"
        assert saved_path.is_file()
        assert metadata["file_size_bytes"] == saved_path.stat().st_size
    finally:
        if saved_path is not None:
            saved_path.unlink(missing_ok=True)


def test_upload_contract_rejects_non_xlsx() -> None:
    app = create_app(settings=Settings(), workflow=ActualDataWorkflow())

    async def request() -> httpx.Response:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test"
        ) as client:
            return await client.post(
                "/v1/comment-korean-column-names",
                files={"file": ("input.csv", b"not,xlsx", "text/csv")},
            )

    response = asyncio.run(request())
    assert response.status_code == 415
