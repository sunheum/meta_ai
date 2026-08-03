from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.excel import RESULT_HEADERS, read_source_columns, write_result_workbook
from app.models import (
    KoreanAttributeResult,
    ProcessingAction,
    ProcessingStatus,
)


ROOT = Path(__file__).resolve().parents[2]
ACTUAL_INPUT = ROOT / "data" / "table_column_template_컬럼코멘트Y.xlsx"


def test_actual_workbook_round_trip_preserves_all_source_rows_and_columns(
) -> None:
    output = Path(__file__).parent / f".result-{uuid4().hex}.xlsx"
    sources = read_source_columns(ACTUAL_INPUT)
    assert len(sources) == 1195
    assert len({(s.column_name, s.column_description) for s in sources}) == 921

    results = []
    for index, source in enumerate(sources):
        review = index == 0
        results.append(
            KoreanAttributeResult(
                source_id=source.source_id,
                original_description=source.column_description,
                korean_attribute_name=source.column_description,
                action=ProcessingAction.KEEP,
                confidence=80 if review else 100,
                reason="",
                review_reasons=["실제 데이터 기반 I/O 검토 표본"] if review else [],
                status=(
                    ProcessingStatus.REVIEW_REQUIRED
                    if review
                    else ProcessingStatus.AUTO_CONFIRMED
                ),
            )
        )

    try:
        write_result_workbook(ACTUAL_INPUT, output, sources, results)

        source_book = load_workbook(ACTUAL_INPUT, data_only=False)
        result_book = load_workbook(output, data_only=False)
        try:
            source_sheet = source_book["테이블_컬럼_정보"]
            result_sheet = result_book["한글속성명_결과"]
            assert result_sheet.max_row == source_sheet.max_row == 1196
            assert result_sheet.max_column == source_sheet.max_column + 6 == 18
            for row_number in range(1, source_sheet.max_row + 1):
                assert [
                    source_sheet.cell(row_number, column).value
                    for column in range(1, 13)
                ] == [
                    result_sheet.cell(row_number, column).value
                    for column in range(1, 13)
                ]
            assert [
                result_sheet.cell(1, column).value for column in range(13, 19)
            ] == list(RESULT_HEADERS)
            assert result_sheet["M2"].value == sources[0].column_description
            assert result_sheet["N2"].value == "검토필요"
            assert result_sheet["O2"].value == 80
            assert result_sheet["A1"].style_id == source_sheet["A1"].style_id

            review_sheet = result_book["검토필요"]
            assert review_sheet.max_row == 2
            assert review_sheet["B2"].value == "row-2"
            assert review_sheet["I2"].value == "검토필요"
        finally:
            source_book.close()
            result_book.close()
    finally:
        output.unlink(missing_ok=True)

    assert not list(Path(__file__).parent.glob(".*.tmp.xlsx"))


def test_actual_workbook_context_fields_are_read() -> None:
    sources = read_source_columns(ACTUAL_INPUT)
    first = sources[0]
    assert first.source_id == "row-2"
    assert first.schema_name
    assert first.table_name
    assert first.table_description
    assert first.column_order is not None
    assert first.data_type
    assert len(first.original_values) == 12
