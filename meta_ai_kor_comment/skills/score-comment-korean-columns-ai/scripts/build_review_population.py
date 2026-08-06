from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


HEADER_ALIASES = {
    "column_name": ("컬럼명", "column_name"),
    "column_description": ("컬럼설명", "column_description"),
    "schema_name": ("스키마", "스키마명", "schema_name"),
    "table_name": ("테이블명", "table_name"),
    "table_description": ("테이블설명", "table_description"),
    "data_type": ("데이터타입", "데이터유형", "data_type"),
    "korean_name": ("한글속성명", "korean_attribute_name"),
    "status": ("처리상태", "processing_status"),
    "confidence": ("신뢰도", "confidence"),
    "method": ("처리방식", "processing_method"),
    "reason": ("변환근거", "conversion_reason"),
    "review_reason": ("검토사유", "review_reason"),
    "source_id": ("source_id", "원본행ID"),
    "risk_codes": ("risk_codes", "위험코드"),
}


@dataclass(frozen=True)
class Table:
    headers: list[str]
    rows: list[dict[str, Any]]


def normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFC", str(value or "")).strip()
    if text.endswith("(*)"):
        text = text[:-3].rstrip()
    return text


def load_table(path: Path, sheet: str | None = None) -> Table:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise RuntimeError("openpyxl is required to read XLSX files") from exc
        workbook = load_workbook(path, read_only=True, data_only=False)
        worksheet = workbook[sheet] if sheet else workbook.active
        values = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(values)
        except StopIteration as exc:
            raise ValueError(f"{path} is empty") from exc
        headers = [str(value or "") for value in raw_headers]
        if len(set(headers)) != len(headers):
            raise ValueError(f"{path} contains duplicate headers")
        rows = [dict(zip(headers, row, strict=False)) for row in values]
        workbook.close()
        return Table(headers, rows)
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = list(reader.fieldnames or [])
            return Table(headers, list(reader))
    text = path.read_text(encoding="utf-8-sig")
    if suffix == ".json":
        value = json.loads(text)
        if isinstance(value, dict):
            value = value.get("rows", value.get("results"))
        if not isinstance(value, list):
            raise ValueError("JSON table must be an array or contain rows/results")
        rows = value
    else:
        rows = [json.loads(line) for line in text.splitlines() if line.strip()]
    if not all(isinstance(row, dict) for row in rows):
        raise ValueError("every table row must be an object")
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return Table(headers, rows)


def header_lookup(headers: Iterable[str]) -> dict[str, str]:
    return {normalize_header(header): header for header in headers}


def field(row: dict[str, Any], lookup: dict[str, str], name: str) -> Any:
    for alias in HEADER_ALIASES[name]:
        actual = lookup.get(normalize_header(alias))
        if actual is not None:
            return row.get(actual)
    return None


def normalized_key(column_name: Any, description: Any) -> tuple[str, str]:
    return (
        unicodedata.normalize("NFC", str(column_name or "")).strip().upper(),
        unicodedata.normalize("NFC", str(description or "")).strip(),
    )


def load_terminology(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    value = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(value, dict):
        value = value.get("terminology_decisions", value.get("groups", []))
    if not isinstance(value, list) or not all(isinstance(item, dict) for item in value):
        raise ValueError("terminology decisions must be an array")
    return value


def parse_risk_codes(value: Any) -> set[str]:
    if isinstance(value, list):
        return {str(item).strip() for item in value if str(item).strip()}
    if not value:
        return set()
    return {part.strip() for part in re.split(r"[,;|]", str(value)) if part.strip()}


def build_population(
    original: Table,
    result: Table,
    terminology: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if len(original.rows) != len(result.rows):
        raise ValueError(
            f"row count mismatch: original={len(original.rows)} result={len(result.rows)}"
        )
    original_lookup = header_lookup(original.headers)
    result_lookup = header_lookup(result.headers)
    for required in ("column_name", "column_description"):
        if field({}, original_lookup, required) is None and not any(
            normalize_header(alias) in original_lookup for alias in HEADER_ALIASES[required]
        ):
            raise ValueError(f"original is missing {required}")
    if not any(
        normalize_header(alias) in result_lookup for alias in HEADER_ALIASES["korean_name"]
    ):
        raise ValueError("result is missing korean_attribute_name")

    pairs: list[tuple[dict[str, Any], dict[str, Any], str, tuple[str, str]]] = []
    key_counts: Counter[tuple[str, str]] = Counter()
    column_descriptions: dict[str, set[str]] = defaultdict(set)
    for index, (source, output) in enumerate(
        zip(original.rows, result.rows, strict=True), start=2
    ):
        column_name = field(source, original_lookup, "column_name")
        description = field(source, original_lookup, "column_description")
        key = normalized_key(column_name, description)
        source_id = str(field(output, result_lookup, "source_id") or f"row-{index}")
        pairs.append((source, output, source_id, key))
        key_counts[key] += 1
        column_descriptions[key[0]].add(key[1])

    decisions_by_source: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for decision in terminology:
        affected = decision.get("affected_source_ids", [])
        if isinstance(affected, list):
            for source_id in affected:
                decisions_by_source[str(source_id)].append(decision)

    seen: set[tuple[str, str]] = set()
    population: list[dict[str, Any]] = []
    for source, output, source_id, key in pairs:
        if key in seen:
            continue
        seen.add(key)
        description = str(field(source, original_lookup, "column_description") or "")
        method = str(field(output, result_lookup, "method") or "")
        status = str(field(output, result_lookup, "status") or "")
        confidence = field(output, result_lookup, "confidence")
        risk_codes = parse_risk_codes(field(output, result_lookup, "risk_codes"))
        strata: set[str] = set()
        if re.search(r"[A-Za-z]", description.replace("ID", "")):
            strata.add("english_translation")
        if re.search(r"[0-9]", description):
            strata.add("numeric_preservation")
        if "/" in description:
            strata.add("slash_context")
        if decisions_by_source.get(source_id) or {
            "terminology_frequency",
            "synonym_frequency",
        } & risk_codes:
            strata.add("terminology_frequency")
        if key_counts[key] > 1 or len(column_descriptions[key[0]]) > 1:
            strata.add("duplicate_context")
        if method in {"재작성", "rewrite"}:
            strata.add("llm_rewrite")
        low_confidence = False
        try:
            low_confidence = int(confidence) < 90
        except (TypeError, ValueError):
            low_confidence = True
        if status not in {"자동확정", "auto_confirmed"} or low_confidence:
            strata.add("review_needed")
        if method in {"유지", "keep"} and not strata:
            strata.add("low_risk_keep")
        if not strata:
            strata.add("other")

        population.append(
            {
                "source_id": source_id,
                "source_ids": [
                    item_source_id
                    for _, _, item_source_id, item_key in pairs
                    if item_key == key
                ],
                "occurrence_count": key_counts[key],
                "스키마": field(source, original_lookup, "schema_name"),
                "테이블명": field(source, original_lookup, "table_name"),
                "테이블설명": field(source, original_lookup, "table_description"),
                "컬럼명": field(source, original_lookup, "column_name"),
                "컬럼설명": field(source, original_lookup, "column_description"),
                "데이터타입": field(source, original_lookup, "data_type"),
                "한글속성명": field(output, result_lookup, "korean_name"),
                "처리상태": field(output, result_lookup, "status"),
                "신뢰도": confidence,
                "처리방식": method,
                "변환근거": field(output, result_lookup, "reason"),
                "검토사유": field(output, result_lookup, "review_reason"),
                "review_strata": sorted(strata),
                "risk_codes": sorted(risk_codes),
                "terminology_decisions": decisions_by_source.get(source_id, []),
            }
        )
    return population


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True))
            handle.write("\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--original", type=Path, required=True)
    parser.add_argument("--result", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--original-sheet")
    parser.add_argument("--result-sheet")
    parser.add_argument("--terminology-decisions", type=Path)
    args = parser.parse_args()
    rows = build_population(
        load_table(args.original, args.original_sheet),
        load_table(args.result, args.result_sheet),
        load_terminology(args.terminology_decisions),
    )
    write_jsonl(args.output, rows)
    print(f"AI review population rows={len(rows)}")


if __name__ == "__main__":
    main()
