from __future__ import annotations

import os
import posixpath
import re
from copy import copy, deepcopy
from pathlib import Path
from typing import Iterable, Sequence
from uuid import uuid4
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exceptions import InputWorkbookError
from app.models import KoreanAttributeResult, ProcessingStatus, SourceColumn

DEFAULT_INPUT_SHEET = "테이블_컬럼_정보"
RESULT_SHEET = "한글속성명_결과"
REVIEW_SHEET = "검토필요"
RESULT_HEADERS = (
    "한글속성명",
    "처리상태",
    "신뢰도",
    "처리방식",
    "변환근거",
    "검토사유",
)
REVIEW_HEADERS = (
    "원본행",
    "source_id",
    "스키마",
    "테이블명",
    "테이블설명",
    "컬럼명",
    "컬럼설명",
    "한글속성명",
    "처리상태",
    "신뢰도",
    "처리방식",
    "변환근거",
    "검토사유",
    "추천확인포인트",
)


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s*\(\*\)\s*$", "", text.strip())
    return re.sub(r"\s+", "", text)


def read_source_columns(
    path: str | Path,
    sheet_name: str = DEFAULT_INPUT_SHEET,
) -> list[SourceColumn]:
    """Read source rows without mutating source values or row order."""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InputWorkbookError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    try:
        worksheet, header_row, header_indexes, header_values = _find_source_sheet(
            workbook, sheet_name
        )
        required = ("컬럼명", "컬럼설명")
        missing = [name for name in required if name not in header_indexes]
        if missing:
            raise InputWorkbookError(
                "필수 헤더를 찾지 못했습니다: " + ", ".join(missing)
            )

        sources: list[SourceColumn] = []
        invalid_rows: list[str] = []
        for excel_row, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if _is_blank_row(row, len(header_values)):
                continue

            column_name = _cell_text(row, header_indexes["컬럼명"], strip=True)
            description = _cell_text(
                row, header_indexes["컬럼설명"], strip=False
            )
            if not column_name or not description.strip():
                invalid_rows.append(
                    f"{excel_row}행(" + (
                        "컬럼명 누락" if not column_name else "컬럼설명 누락"
                    ) + ")"
                )
                continue

            original_values = {
                normalized: row[index] if index < len(row) else None
                for index, normalized in enumerate(header_values)
                if normalized
            }
            sources.append(
                SourceColumn(
                    source_id=f"row-{excel_row}",
                    column_name=column_name.upper(),
                    column_description=description,
                    schema_name=_optional_text(row, header_indexes.get("스키마")),
                    table_name=_optional_text(row, header_indexes.get("테이블명")),
                    table_description=_optional_text(
                        row, header_indexes.get("테이블설명")
                    ),
                    column_order=_optional_raw(
                        row, header_indexes.get("컬럼순번")
                    ),
                    data_type=_optional_text(
                        row, header_indexes.get("데이터타입")
                    ),
                    original_values=original_values,
                )
            )

        if invalid_rows:
            preview = ", ".join(invalid_rows[:10])
            suffix = "" if len(invalid_rows) <= 10 else f" 외 {len(invalid_rows) - 10}건"
            raise InputWorkbookError(
                f"컬럼명과 컬럼설명은 필수입니다: {preview}{suffix}"
            )
        if not sources:
            raise InputWorkbookError("처리할 컬럼 데이터가 없습니다.")
        return sources
    finally:
        workbook.close()


def write_result_workbook(
    input_path: str | Path,
    output_path: str | Path,
    sources: Sequence[SourceColumn],
    results: Iterable[KoreanAttributeResult],
    *,
    input_sheet_name: str = DEFAULT_INPUT_SHEET,
    result_sheet_name: str = RESULT_SHEET,
    review_sheet_name: str = REVIEW_SHEET,
) -> None:
    """Add result columns to a copy of the input workbook and save atomically."""

    source_rows = list(sources)
    result_rows = list(results)
    if not source_rows:
        raise InputWorkbookError("출력할 원본 행이 없습니다.")
    result_by_id = _index_results(result_rows)
    source_ids = [source.source_id for source in source_rows]
    missing = [source_id for source_id in source_ids if source_id not in result_by_id]
    extras = sorted(set(result_by_id).difference(source_ids))
    if missing or extras:
        messages: list[str] = []
        if missing:
            messages.append(f"결과 누락 {len(missing)}건")
        if extras:
            messages.append(f"원본에 없는 결과 {len(extras)}건")
        raise InputWorkbookError("원본-결과 대응 오류: " + ", ".join(messages))

    try:
        workbook = load_workbook(input_path, data_only=False)
    except Exception as exc:
        raise InputWorkbookError(f"출력용 엑셀 파일을 열 수 없습니다: {exc}") from exc

    try:
        worksheet, header_row, _, header_values = _find_source_sheet(
            workbook, input_sheet_name
        )
        original_sheet_title = worksheet.title
        if result_sheet_name in workbook.sheetnames and worksheet.title != result_sheet_name:
            workbook.remove(workbook[result_sheet_name])
        worksheet.title = result_sheet_name

        last_original_column = max(
            index for index, value in enumerate(header_values, start=1) if value
        )
        _clear_previous_results(
            worksheet, header_row, last_original_column, len(RESULT_HEADERS)
        )
        _write_result_columns(
            worksheet,
            header_row,
            last_original_column,
            source_rows,
            result_by_id,
        )
        _expand_existing_tables(
            worksheet,
            header_row=header_row,
            last_column=last_original_column + len(RESULT_HEADERS),
        )

        if review_sheet_name in workbook.sheetnames:
            workbook.remove(workbook[review_sheet_name])
        review_sheet = workbook.create_sheet(review_sheet_name)
        review_rows = [
            (source, result_by_id[source.source_id])
            for source in source_rows
            if result_by_id[source.source_id].status
            in {
                ProcessingStatus.REVIEW_REQUIRED,
                ProcessingStatus.VALIDATION_FAILED,
            }
        ]
        _write_review_sheet(review_sheet, review_rows)
        _save_workbook_atomic(
            workbook,
            Path(output_path),
            source_path=Path(input_path),
            source_sheet_name=original_sheet_title,
            result_sheet_name=result_sheet_name,
            original_column_count=last_original_column,
        )
    finally:
        workbook.close()


def _find_source_sheet(workbook, preferred_name: str):
    candidates = []
    if preferred_name in workbook.sheetnames:
        candidates.append(workbook[preferred_name])
    candidates.extend(
        sheet for sheet in workbook.worksheets if sheet not in candidates
    )
    for worksheet in candidates:
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(20, worksheet.max_row),
                values_only=True,
            ),
            start=1,
        ):
            headers = [_normalize_header(value) for value in row]
            indexes = {
                header: index
                for index, header in enumerate(headers)
                if header
            }
            if {"컬럼명", "컬럼설명"}.issubset(indexes):
                return worksheet, row_number, indexes, headers
    raise InputWorkbookError(
        "헤더에서 '컬럼명'과 '컬럼설명'을 찾지 못했습니다. "
        "'컬럼명 (*)' 표기도 허용됩니다."
    )


def _write_result_columns(
    worksheet,
    header_row: int,
    last_original_column: int,
    sources: Sequence[SourceColumn],
    result_by_id: dict[str, KoreanAttributeResult],
) -> None:
    first_result_column = last_original_column + 1
    source_header = worksheet.cell(header_row, last_original_column)
    for offset, header in enumerate(RESULT_HEADERS):
        cell = worksheet.cell(header_row, first_result_column + offset, header)
        if source_header.has_style:
            cell._style = copy(source_header._style)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for source in sources:
        excel_row = _excel_row(source.source_id)
        result = result_by_id[source.source_id]
        review_reason = " | ".join(result.review_reasons)
        terminology_reason = " | ".join(result.terminology_decisions)
        reason = result.reason.strip()
        if terminology_reason:
            reason = " | ".join(part for part in (reason, terminology_reason) if part)
        values = (
            result.korean_attribute_name,
            result.status.value,
            result.confidence,
            result.processing_method,
            reason,
            review_reason,
        )
        source_style_cell = worksheet.cell(excel_row, last_original_column)
        for offset, value in enumerate(values):
            cell = worksheet.cell(excel_row, first_result_column + offset, value)
            if source_style_cell.has_style:
                cell._style = copy(source_style_cell._style)
            cell.alignment = Alignment(
                horizontal="right" if offset == 2 else "left",
                vertical="top",
                wrap_text=offset in {4, 5},
            )
        worksheet.cell(excel_row, first_result_column + 2).number_format = "0"

    widths = (28, 14, 10, 12, 56, 56)
    for offset, width in enumerate(widths):
        worksheet.column_dimensions[
            get_column_letter(first_result_column + offset)
        ].width = width
    worksheet.freeze_panes = worksheet.freeze_panes or f"A{header_row + 1}"
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(first_result_column + len(RESULT_HEADERS) - 1)}"
        f"{worksheet.max_row}"
    )


def _write_review_sheet(worksheet, rows) -> None:
    worksheet.title = REVIEW_SHEET if worksheet.title == "Sheet" else worksheet.title
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = "C00000"
    worksheet.freeze_panes = "A2"
    worksheet.append(list(REVIEW_HEADERS))

    for source, result in rows:
        review_reason = " | ".join(result.review_reasons)
        terminology_reason = " | ".join(result.terminology_decisions)
        reason = " | ".join(
            part for part in (result.reason.strip(), terminology_reason) if part
        )
        recommended = _recommended_review_point(result)
        worksheet.append(
            [
                _excel_row(source.source_id),
                source.source_id,
                source.schema_name,
                source.table_name,
                source.table_description,
                source.column_name,
                source.column_description,
                result.korean_attribute_name,
                result.status.value,
                result.confidence,
                result.processing_method,
                reason,
                review_reason,
                recommended,
            ]
        )

    _style_header(worksheet, "C00000")
    widths = (10, 14, 14, 24, 32, 26, 32, 28, 14, 10, 12, 56, 56, 48)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2, max_col=len(REVIEW_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="right" if cell.column in {1, 10} else "left",
                vertical="top",
                wrap_text=cell.column in {5, 7, 12, 13, 14},
            )
    if rows:
        table = Table(
            displayName="KoreanAttributeReview",
            ref=f"A1:{get_column_letter(len(REVIEW_HEADERS))}{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium3",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(REVIEW_HEADERS))}{max(1, len(rows) + 1)}"
    )


def _recommended_review_point(result: KoreanAttributeResult) -> str:
    parts: list[str] = []
    if result.validation_issue_codes:
        parts.append("검증코드 확인: " + ", ".join(result.validation_issue_codes))
    if result.review_reasons:
        parts.append("문맥·업무 의미 확인")
    if result.status is ProcessingStatus.VALIDATION_FAILED:
        parts.append("결정적 정책 위반 수정 필요")
    return " | ".join(parts) or "생성명과 원문 의미 확인"


def _clear_previous_results(
    worksheet, header_row: int, last_original_column: int, result_width: int
) -> None:
    start = last_original_column + 1
    end = start + result_width - 1
    for row in worksheet.iter_rows(
        min_row=header_row,
        max_row=worksheet.max_row,
        min_col=start,
        max_col=end,
    ):
        for cell in row:
            cell.value = None


def _expand_existing_tables(worksheet, header_row: int, last_column: int) -> None:
    for table in worksheet.tables.values():
        min_col, min_row, _, max_row = range_boundaries(table.ref)
        if min_row == header_row:
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(last_column)}{max_row}"
            )


def _save_workbook_atomic(
    workbook: Workbook,
    path: Path,
    *,
    source_path: Path,
    source_sheet_name: str,
    result_sheet_name: str,
    original_column_count: int,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    staging = path.with_name(f".{path.stem}.{uuid4().hex}.tmp{path.suffix}")
    try:
        workbook.save(staging)
        _restore_original_cell_xml(
            source_path,
            staging,
            source_sheet_name=source_sheet_name,
            result_sheet_name=result_sheet_name,
            original_column_count=original_column_count,
        )
        os.replace(staging, path)
    except Exception as exc:
        raise InputWorkbookError(f"결과 엑셀 파일을 저장할 수 없습니다: {exc}") from exc
    finally:
        staging.unlink(missing_ok=True)


def _restore_original_cell_xml(
    source_path: Path,
    output_path: Path,
    *,
    source_sheet_name: str,
    result_sheet_name: str,
    original_column_count: int,
) -> None:
    """Restore original cell nodes and shared strings for exact A:L values.

    openpyxl serializes an empty shared string as an empty inline-string cell,
    which reloads as ``None`` instead of ``""``. Transplanting the original cell
    nodes keeps source values and style ids exact while retaining generated cells.
    """

    patched_path = output_path.with_name(
        f".{output_path.stem}.{uuid4().hex}.patched{output_path.suffix}"
    )
    try:
        with ZipFile(source_path, "r") as source_zip, ZipFile(
            output_path, "r"
        ) as output_zip:
            source_sheet_part = _worksheet_part(source_zip, source_sheet_name)
            result_sheet_part = _worksheet_part(output_zip, result_sheet_name)
            source_tree = ET.fromstring(source_zip.read(source_sheet_part))
            result_tree = ET.fromstring(output_zip.read(result_sheet_part))
            _transplant_cells(
                source_tree,
                result_tree,
                original_column_count=original_column_count,
            )

            replacements: dict[str, bytes] = {
                result_sheet_part: ET.tostring(
                    result_tree, encoding="utf-8", xml_declaration=True
                )
            }
            shared_part = "xl/sharedStrings.xml"
            if shared_part in source_zip.namelist():
                replacements[shared_part] = source_zip.read(shared_part)
                replacements.update(_shared_string_package_updates(output_zip))

            with ZipFile(patched_path, "w", compression=ZIP_DEFLATED) as patched_zip:
                replaced = set(replacements)
                for info in output_zip.infolist():
                    if info.filename not in replaced:
                        patched_zip.writestr(info, output_zip.read(info.filename))
                for name, data in replacements.items():
                    patched_zip.writestr(name, data)
        os.replace(patched_path, output_path)
    except Exception as exc:
        raise InputWorkbookError(
            f"원본 셀 무결성을 복원할 수 없습니다: {exc}"
        ) from exc
    finally:
        patched_path.unlink(missing_ok=True)


def _worksheet_part(archive: ZipFile, sheet_name: str) -> str:
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = next(
        (
            node
            for node in workbook.findall(f".//{{{spreadsheet_ns}}}sheet")
            if node.attrib.get("name") == sheet_name
        ),
        None,
    )
    if sheet is None:
        raise InputWorkbookError(f"워크시트 관계를 찾을 수 없습니다: {sheet_name}")
    relation_id = sheet.attrib[f"{{{office_rel_ns}}}id"]
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relation = next(
        (
            node
            for node in relationships.findall(f"{{{package_rel_ns}}}Relationship")
            if node.attrib.get("Id") == relation_id
        ),
        None,
    )
    if relation is None:
        raise InputWorkbookError(f"워크시트 파일 관계가 없습니다: {sheet_name}")
    target = relation.attrib["Target"].replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _transplant_cells(source_tree, result_tree, *, original_column_count: int) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    row_tag = f"{{{namespace}}}row"
    cell_tag = f"{{{namespace}}}c"
    source_data = source_tree.find(f"{{{namespace}}}sheetData")
    result_data = result_tree.find(f"{{{namespace}}}sheetData")
    if source_data is None or result_data is None:
        raise InputWorkbookError("워크시트 sheetData를 찾을 수 없습니다.")
    source_rows = {row.attrib.get("r"): row for row in source_data.findall(row_tag)}
    result_rows = {row.attrib.get("r"): row for row in result_data.findall(row_tag)}
    for row_id, source_row in source_rows.items():
        result_row = result_rows.get(row_id)
        if result_row is None:
            result_row = deepcopy(source_row)
            result_data.append(result_row)
            continue
        original_cells = [
            deepcopy(cell)
            for cell in source_row.findall(cell_tag)
            if _cell_column(cell.attrib.get("r", "")) <= original_column_count
        ]
        generated_cells = [
            deepcopy(cell)
            for cell in result_row.findall(cell_tag)
            if _cell_column(cell.attrib.get("r", "")) > original_column_count
        ]
        for cell in list(result_row.findall(cell_tag)):
            result_row.remove(cell)
        for index, cell in enumerate(original_cells + generated_cells):
            result_row.insert(index, cell)


def _cell_column(reference: str) -> int:
    match = re.match(r"([A-Z]+)", reference.upper())
    if match is None:
        return 0
    return column_index_from_string(match.group(1))


def _shared_string_package_updates(output_zip: ZipFile) -> dict[str, bytes]:
    relationships_path = "xl/_rels/workbook.xml.rels"
    content_types_path = "[Content_Types].xml"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    shared_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        "sharedStrings"
    )
    shared_content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml."
        "sharedStrings+xml"
    )

    relationships = ET.fromstring(output_zip.read(relationships_path))
    relation_tag = f"{{{package_rel_ns}}}Relationship"
    if not any(
        node.attrib.get("Type") == shared_type
        for node in relationships.findall(relation_tag)
    ):
        used_ids = {
            int(match.group(1))
            for node in relationships.findall(relation_tag)
            if (match := re.fullmatch(r"rId(\d+)", node.attrib.get("Id", "")))
        }
        next_id = max(used_ids, default=0) + 1
        ET.SubElement(
            relationships,
            relation_tag,
            {
                "Type": shared_type,
                "Target": "sharedStrings.xml",
                "Id": f"rId{next_id}",
            },
        )

    content_types = ET.fromstring(output_zip.read(content_types_path))
    override_tag = f"{{{content_ns}}}Override"
    if not any(
        node.attrib.get("PartName") == "/xl/sharedStrings.xml"
        for node in content_types.findall(override_tag)
    ):
        ET.SubElement(
            content_types,
            override_tag,
            {
                "PartName": "/xl/sharedStrings.xml",
                "ContentType": shared_content_type,
            },
        )

    return {
        relationships_path: ET.tostring(
            relationships, encoding="utf-8", xml_declaration=True
        ),
        content_types_path: ET.tostring(
            content_types, encoding="utf-8", xml_declaration=True
        ),
    }


def _index_results(
    results: Sequence[KoreanAttributeResult],
) -> dict[str, KoreanAttributeResult]:
    indexed: dict[str, KoreanAttributeResult] = {}
    duplicates: list[str] = []
    for result in results:
        if result.source_id in indexed:
            duplicates.append(result.source_id)
        indexed[result.source_id] = result
    if duplicates:
        raise InputWorkbookError(
            "중복 결과 source_id가 있습니다: " + ", ".join(sorted(set(duplicates)))
        )
    return indexed


def _style_header(worksheet, fill_color: str) -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28


def _excel_row(source_id: str) -> int:
    match = re.fullmatch(r"row-(\d+)", source_id)
    if match is None:
        raise InputWorkbookError(f"잘못된 source_id 형식입니다: {source_id}")
    return int(match.group(1))


def _is_blank_row(row: tuple[object, ...], header_width: int) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row[:header_width]
    )


def _cell_text(row: tuple[object, ...], index: int, *, strip: bool) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    value = str(row[index])
    return value.strip() if strip else value


def _optional_text(row: tuple[object, ...], index: int | None) -> str | None:
    if index is None:
        return None
    value = _cell_text(row, index, strip=True)
    return value or None


def _optional_raw(row: tuple[object, ...], index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]
