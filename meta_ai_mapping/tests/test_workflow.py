import asyncio
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook

from app.glossary import CanonicalGlossary, canonical_key
from app.models import (
    MappingCandidate,
    ProgressEvent,
    SourceColumn,
    ValidationIssue,
    WorkflowOptions,
)
from app.workflow import MappingWorkflow, reconcile_mappings


class FakeReviewModel:
    def __init__(self) -> None:
        self.review_calls = 0

    async def generate(
        self, sources: Sequence[SourceColumn]
    ) -> list[MappingCandidate]:
        return [
            MappingCandidate(
                source_id=sources[0].source_id,
                abbreviation="BAD",
                full_name="DATE",
                korean_word="일자",
            )
        ]

    async def review(
        self,
        sources: Sequence[SourceColumn],
        current_mappings: Sequence[MappingCandidate],
        issues: Sequence[ValidationIssue],
        review_round: int,
    ) -> list[MappingCandidate]:
        self.review_calls += 1
        return [
            MappingCandidate(
                source_id=sources[0].source_id,
                abbreviation="DT",
                full_name="DATE",
                korean_word="일자",
            )
        ]


class PartialFailureModel:
    async def generate(
        self, sources: Sequence[SourceColumn]
    ) -> list[MappingCandidate]:
        return [
            MappingCandidate(
                source_id=sources[0].source_id,
                abbreviation="DT",
                full_name="DATE",
                korean_word="일자",
            ),
            MappingCandidate(
                source_id=sources[1].source_id,
                abbreviation="BAD",
                full_name="BAD",
                korean_word="년",
            ),
        ]

    async def review(
        self,
        sources: Sequence[SourceColumn],
        current_mappings: Sequence[MappingCandidate],
        issues: Sequence[ValidationIssue],
        review_round: int,
    ) -> list[MappingCandidate]:
        return list(current_mappings)


class DuplicateSourceModel:
    def __init__(self) -> None:
        self.generated_source_counts: list[int] = []

    async def generate(
        self, sources: Sequence[SourceColumn]
    ) -> list[MappingCandidate]:
        self.generated_source_counts.append(len(sources))
        return [
            MappingCandidate(
                source_id=source.source_id,
                abbreviation="AP",
                full_name="APPLY",
                korean_word="적용",
            )
            for source in sources
        ]

    async def review(
        self,
        sources: Sequence[SourceColumn],
        current_mappings: Sequence[MappingCandidate],
        issues: Sequence[ValidationIssue],
        review_round: int,
    ) -> list[MappingCandidate]:
        raise AssertionError("전역 표준화 후에는 리뷰가 필요하지 않습니다.")


def test_workflow_reviews_and_fixes_invalid_mapping(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["컬럼명 (*)", "컬럼설명"])
    sheet.append(["STDT", "기준일자"])
    workbook.save(input_path)

    model = FakeReviewModel()
    workflow = MappingWorkflow(model)
    events: list[ProgressEvent] = []

    async def execute():
        async def capture(event: ProgressEvent) -> None:
            events.append(event)

        return await workflow.run(
            input_path,
            output_path,
            WorkflowOptions(batch_size=10, max_concurrency=10, max_review_rounds=2),
            progress_callback=capture,
        )

    result = asyncio.run(execute())

    assert result.validation_report.is_valid
    assert result.review_rounds == 1
    assert model.review_calls == 1
    output = load_workbook(output_path, read_only=True).active
    assert output["A2"].value == "DT"
    assert output["D2"].value == 1
    assert [event.stage for event in events] == [
        "input",
        "input",
        "generate",
        "generate",
        "reconcile",
        "reconcile",
        "validate",
        "validate",
        "review",
        "review",
        "reconcile",
        "reconcile",
        "validate",
        "validate",
        "output",
        "output",
    ]
    assert events[-1].overall_percent == 100


def test_workflow_writes_partial_failures_to_separate_sheet(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "partial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["컬럼명 (*)", "컬럼설명"])
    sheet.append(["STDT", "기준일자"])
    sheet.append(["YYMM", "년월"])
    workbook.save(input_path)

    result = asyncio.run(
        MappingWorkflow(PartialFailureModel()).run(
            input_path,
            output_path,
            WorkflowOptions(
                batch_size=10,
                max_concurrency=10,
                max_review_rounds=0,
            ),
        )
    )

    assert result.is_partial
    assert result.failed_source_count == 1
    assert not result.validation_report.is_valid
    output = load_workbook(output_path, read_only=True)
    assert output.sheetnames == ["약어_매핑", "검증실패"]
    assert output["약어_매핑"]["A2"].value == "DT"
    assert output["검증실패"]["A2"].value == "row-3"
    assert output["검증실패"]["F2"].value == "BAD"
    assert "abbreviation_not_in_column" in output["검증실패"]["I2"].value
    assert "YYMM" in output["검증실패"]["K2"].value


def test_workflow_generates_duplicate_sources_once_and_applies_glossary(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "duplicates.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["테이블명", "컬럼명 (*)", "컬럼설명"])
    sheet.append(["TABLE_A", "AP_DT", "적용일자"])
    sheet.append(["TABLE_B", "AP_DT", "적용일자"])
    workbook.save(input_path)

    model = DuplicateSourceModel()
    glossary = CanonicalGlossary(
        entries={canonical_key("AP", "적용"): "APPLICATION"}
    )
    result = asyncio.run(
        MappingWorkflow(model, glossary=glossary).run(
            input_path,
            output_path,
            WorkflowOptions(
                batch_size=10,
                max_concurrency=10,
                max_review_rounds=2,
            ),
        )
    )

    assert result.validation_report.is_valid
    assert model.generated_source_counts == [1]
    output = load_workbook(output_path, read_only=True).active
    assert output["A2"].value == "AP"
    assert output["B2"].value == "APPLICATION"
    assert output["C2"].value == "적용"
    assert output["D2"].value == 2
    assert result.reconciliation_stats["glossary_replacement_count"] == 2


def test_reconcile_uses_majority_for_unknown_mapping() -> None:
    candidates = [
        MappingCandidate(
            source_id="row-2",
            abbreviation="ABC",
            full_name="ALPHA",
            korean_word="항목",
        ),
        MappingCandidate(
            source_id="row-3",
            abbreviation="ABC",
            full_name="ALPHA",
            korean_word="항목",
        ),
        MappingCandidate(
            source_id="row-4",
            abbreviation="ABC",
            full_name="ALPHABET",
            korean_word="항목",
        ),
    ]

    reconciled, stats = reconcile_mappings(
        candidates,
        CanonicalGlossary.empty(),
    )

    assert {candidate.full_name for candidate in reconciled} == {"ALPHA"}
    assert stats["majority_replacement_count"] == 1
