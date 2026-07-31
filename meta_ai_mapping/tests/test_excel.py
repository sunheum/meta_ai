from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.excel import read_source_columns, write_mapping_workbook
from app.models import MappingSummary


def _make_input(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["테이블명 (*)", "컬럼명 (*)", "컬럼설명"])
    sheet.append(["COM_CLD", "STDT", "기준일자"])
    sheet.append(["COM_CLD", "YYMM", "년월"])
    workbook.save(path)


def test_read_source_columns_and_write_output(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_input(input_path)

    sources = read_source_columns(input_path)

    assert [source.source_id for source in sources] == ["row-2", "row-3"]
    assert sources[0].column_name == "STDT"
    assert sources[0].column_description == "기준일자"

    write_mapping_workbook(
        output_path,
        [
            MappingSummary(
                abbreviation="DT",
                full_name="DATE",
                korean_word="일자",
                occurrence_count=2,
            )
        ],
    )
    workbook = load_workbook(output_path, read_only=True)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        "영문약어",
        "영문 Full Name",
        "한글단어",
        "출현건수",
    ]
    assert [cell.value for cell in sheet[2]] == ["DT", "DATE", "일자", 2]

