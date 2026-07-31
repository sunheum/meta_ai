import asyncio
from io import BytesIO
from pathlib import Path
from typing import Sequence

import httpx
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.main import create_app
from app.models import (
    MappingCandidate,
    SourceColumn,
    ValidationIssue,
)
from app.workflow import MappingWorkflow


class FakeApiModel:
    def __init__(self, delay_seconds: float = 0) -> None:
        self.delay_seconds = delay_seconds

    async def generate(
        self, sources: Sequence[SourceColumn]
    ) -> list[MappingCandidate]:
        if self.delay_seconds:
            await asyncio.sleep(self.delay_seconds)
        return [
            MappingCandidate(
                source_id=source.source_id,
                abbreviation="DT",
                full_name="DATE",
                korean_word="일자",
            )
            for source in sources
        ]

    async def review(
        self,
        sources: Sequence[SourceColumn],
        current_mappings: Sequence[MappingCandidate],
        issues: Sequence[ValidationIssue],
        review_round: int,
    ) -> list[MappingCandidate]:
        raise AssertionError("유효한 생성 결과에는 리뷰가 호출되면 안 됩니다.")


class PartialApiModel(FakeApiModel):
    async def generate(
        self, sources: Sequence[SourceColumn]
    ) -> list[MappingCandidate]:
        return [
            MappingCandidate(
                source_id=sources[0].source_id,
                abbreviation="DT",
                full_name="DATE",
                korean_word="일자",
            ),
            MappingCandidate(
                source_id=sources[1].source_id,
                abbreviation="BAD",
                full_name="BAD",
                korean_word="년",
            ),
        ]


def _input_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["컬럼명 (*)", "컬럼설명"])
    sheet.append(["STDT", "기준일자"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _partial_input_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["컬럼명 (*)", "컬럼설명"])
    sheet.append(["STDT", "기준일자"])
    sheet.append(["YYMM", "년월"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_api_returns_xlsx() -> None:
    app = create_app(
        settings=Settings(),
        workflow=MappingWorkflow(FakeApiModel()),
    )

    async def request() -> httpx.Response:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test"
        ) as client:
            return await client.post(
                "/v1/abbreviation-mappings",
                files={
                    "file": (
                        "table_column_template.xlsx",
                        _input_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

    response = asyncio.run(request())

    assert response.status_code == 200
    assert response.headers["x-source-count"] == "1"
    assert response.headers["x-mapping-count"] == "1"
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.active["A2"].value == "DT"


def test_job_api_streams_progress_and_saves_xlsx(tmp_path: Path) -> None:
    results_dir = tmp_path / "results"
    app = create_app(
        settings=Settings(
            progress_heartbeat_seconds=0.01,
            results_dir=str(results_dir),
        ),
        workflow=MappingWorkflow(FakeApiModel(delay_seconds=0.05)),
    )

    async def request() -> tuple[
        httpx.Response,
        httpx.Response,
        httpx.Response,
        httpx.Response,
    ]:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test"
        ) as client:
            created = await client.post(
                "/v1/abbreviation-mappings/jobs",
                files={
                    "file": (
                        "table_column_template.xlsx",
                        _input_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            job_id = created.json()["job_id"]
            progress = await client.get(
                f"/v1/abbreviation-mappings/jobs/{job_id}/progress"
            )
            status = await client.get(
                f"/v1/abbreviation-mappings/jobs/{job_id}"
            )
            removed_download = await client.get(
                f"/v1/abbreviation-mappings/jobs/{job_id}/result"
            )
            return created, progress, status, removed_download

    created, progress, status, removed_download = asyncio.run(request())

    assert created.status_code == 202
    assert "curl_progress" in created.json()
    assert "result_url" not in created.json()
    assert created.json()["result_path"].endswith(".xlsx")
    assert progress.status_code == 200
    assert "LLM 생성" in progress.text
    assert "전역 표준화" in progress.text
    assert "검증 통과" in progress.text
    assert "████" in progress.text
    assert "연결 유지" in progress.text
    assert "XLSX 생성 완료" in progress.text
    assert "단계 누적" in progress.text
    assert "전체 소요" in progress.text
    assert "단계별 소요시간" in progress.text
    assert status.status_code == 200
    timing = status.json()["timing"]
    assert timing["total_elapsed_seconds"] > 0
    assert timing["stage_durations_seconds"]["generate"] > 0
    assert "validate" in timing["stage_durations_seconds"]
    metadata = status.json()["result_metadata"]
    saved_path = Path(metadata["result_path"])
    assert saved_path == results_dir / f"{created.json()['job_id']}.xlsx"
    assert metadata["file_size_bytes"] == saved_path.stat().st_size
    assert "결과 저장 완료" in progress.text
    assert saved_path.is_file()
    assert removed_download.status_code == 404
    workbook = load_workbook(saved_path, read_only=True)
    assert workbook.active["A2"].value == "DT"


def test_job_api_saves_partial_result(tmp_path: Path) -> None:
    results_dir = tmp_path / "results"
    app = create_app(
        settings=Settings(results_dir=str(results_dir)),
        workflow=MappingWorkflow(PartialApiModel()),
    )

    async def request() -> tuple[httpx.Response, httpx.Response]:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test"
        ) as client:
            created = await client.post(
                "/v1/abbreviation-mappings/jobs",
                files={
                    "file": (
                        "table_column_template.xlsx",
                        _partial_input_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"max_review_rounds": "0"},
            )
            job_id = created.json()["job_id"]
            await client.get(
                f"/v1/abbreviation-mappings/jobs/{job_id}/progress"
            )
            status = await client.get(
                f"/v1/abbreviation-mappings/jobs/{job_id}"
            )
            return created, status

    created, status = asyncio.run(request())

    assert created.status_code == 202
    assert status.status_code == 200
    metadata = status.json()["result_metadata"]
    assert metadata["is_partial"] is True
    assert metadata["failed_source_count"] == 1
    saved_path = Path(metadata["result_path"])
    assert saved_path == results_dir / f"{created.json()['job_id']}.xlsx"
    workbook = load_workbook(saved_path, read_only=True)
    assert workbook.sheetnames == ["약어_매핑", "검증실패"]
