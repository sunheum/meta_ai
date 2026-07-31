from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(frozen=True, slots=True)
class CanonicalGlossary:
    entries: dict[tuple[str, str], str] = field(default_factory=dict)
    source_path: str | None = None

    @classmethod
    def empty(cls) -> "CanonicalGlossary":
        return cls()

    @classmethod
    def from_xlsx(cls, path: str | Path) -> "CanonicalGlossary":
        source_path = Path(path)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row, headers = _find_header(worksheet)
        abbreviation_index = headers["영문약어"]
        full_name_index = headers["영문FullName"]
        korean_word_index = headers["한글단어"]
        occurrence_index = headers.get("출현건수")
        selected: dict[tuple[str, str], tuple[str, int]] = {}

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            abbreviation = _cell_text(row, abbreviation_index).upper()
            full_name = _cell_text(row, full_name_index).upper()
            korean_word = _cell_text(row, korean_word_index)
            if not abbreviation or not full_name or not korean_word:
                continue
            occurrence_count = _cell_int(row, occurrence_index)
            key = canonical_key(abbreviation, korean_word)
            previous = selected.get(key)
            if previous is None or occurrence_count > previous[1]:
                selected[key] = (full_name, occurrence_count)

        workbook.close()
        return cls(
            entries={key: value[0] for key, value in selected.items()},
            source_path=str(source_path),
        )

    def get(self, abbreviation: str, korean_word: str) -> str | None:
        return self.entries.get(canonical_key(abbreviation, korean_word))

    def __len__(self) -> int:
        return len(self.entries)


def load_canonical_glossary(path: str | Path | None) -> CanonicalGlossary:
    if not path:
        return CanonicalGlossary.empty()
    source_path = Path(path)
    if not source_path.is_file():
        logger.warning(
            "표준 약어 사전 파일을 찾지 못해 빈 사전으로 시작합니다: %s",
            source_path,
        )
        return CanonicalGlossary.empty()
    try:
        glossary = CanonicalGlossary.from_xlsx(source_path)
    except Exception:
        logger.exception(
            "표준 약어 사전을 읽지 못해 빈 사전으로 시작합니다: %s",
            source_path,
        )
        return CanonicalGlossary.empty()
    logger.info(
        "표준 약어 사전 %d건을 로드했습니다: %s",
        len(glossary),
        source_path,
    )
    return glossary


def canonical_key(abbreviation: str, korean_word: str) -> tuple[str, str]:
    normalized_korean = re.sub(r"[\s_\-/·(),.]", "", korean_word.strip())
    return abbreviation.strip().upper(), normalized_korean


def _find_header(worksheet) -> tuple[int, dict[str, int]]:
    wanted = {"영문약어", "영문FullName", "한글단어"}
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(30, worksheet.max_row),
            values_only=True,
        ),
        start=1,
    ):
        headers = {
            _normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None
        }
        if wanted.issubset(headers):
            return row_number, headers
    raise ValueError(
        "표준 사전에서 '영문약어', '영문 Full Name', '한글단어' 헤더를 "
        "찾지 못했습니다."
    )


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value).strip())


def _cell_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _cell_int(row: tuple[object, ...], index: int | None) -> int:
    if index is None or index >= len(row) or row[index] is None:
        return 0
    try:
        return int(row[index])
    except (TypeError, ValueError):
        return 0
