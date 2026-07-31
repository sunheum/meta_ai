from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exceptions import InputWorkbookError
from app.models import FailedMappingRow, MappingSummary, SourceColumn


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s*\(\*\)\s*$", "", text.strip())
    return re.sub(r"\s+", "", text)


def read_source_columns(path: str | Path) -> list[SourceColumn]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InputWorkbookError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    worksheet = workbook.active
    header_row: int | None = None
    header_indexes: dict[str, int] = {}
    wanted = {"컬럼명", "컬럼설명"}

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True),
        start=1,
    ):
        indexes = {
            _normalize_header(value): index
            for index, value in enumerate(row)
            if value is not None
        }
        if wanted.issubset(indexes):
            header_row = row_number
            header_indexes = indexes
            break

    if header_row is None:
        workbook.close()
        raise InputWorkbookError(
            "헤더에서 '컬럼명'과 '컬럼설명'을 찾지 못했습니다. "
            "'컬럼명 (*)' 표기도 허용됩니다."
        )

    column_name_index = header_indexes["컬럼명"]
    description_index = header_indexes["컬럼설명"]
    table_index = header_indexes.get("테이블명")
    schema_index = header_indexes.get("스키마")
    sources: list[SourceColumn] = []

    for excel_row, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        column_name = _cell_text(row, column_name_index)
        description = _cell_text(row, description_index)
        if not column_name and not description:
            continue
        if not column_name:
            continue
        sources.append(
            SourceColumn(
                source_id=f"row-{excel_row}",
                column_name=column_name.upper(),
                column_description=description,
                table_name=_optional_cell_text(row, table_index),
                schema_name=_optional_cell_text(row, schema_index),
            )
        )

    workbook.close()
    if not sources:
        raise InputWorkbookError("처리할 컬럼 데이터가 없습니다.")
    return sources


def write_mapping_workbook(
    path: str | Path,
    summaries: Iterable[MappingSummary],
    failed_mappings: Iterable[FailedMappingRow] | None = None,
) -> None:
    rows = list(summaries)
    failed_rows = list(failed_mappings or [])
    if not rows and not failed_rows:
        raise InputWorkbookError("출력할 매핑 결과가 없습니다.")

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "약어_매핑"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    headers = ["영문약어", "영문 Full Name", "한글단어", "출현건수"]
    worksheet.append(headers)
    for summary in rows:
        worksheet.append(
            [
                summary.abbreviation,
                summary.full_name,
                summary.korean_word,
                summary.occurrence_count,
            ]
        )

    _style_header(worksheet, "4472C4")

    widths = {"A": 18, "B": 34, "C": 24, "D": 14}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2, max_col=4):
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        row[1].alignment = Alignment(horizontal="left", vertical="center")
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[3].alignment = Alignment(horizontal="right", vertical="center")
        row[3].number_format = "#,##0"

    if rows:
        _add_table(
            worksheet,
            display_name="AbbreviationMappings",
            reference=f"A1:D{len(rows) + 1}",
            style_name="TableStyleMedium2",
        )
    worksheet.auto_filter.ref = f"A1:D{len(rows) + 1}"

    if failed_rows:
        _add_failed_mapping_sheet(workbook, failed_rows)

    workbook.save(output_path)


def _add_failed_mapping_sheet(
    workbook: Workbook, failed_rows: list[FailedMappingRow]
) -> None:
    worksheet = workbook.create_sheet("검증실패")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = "C00000"
    worksheet.freeze_panes = "A2"
    headers = [
        "원본행",
        "스키마",
        "테이블명",
        "컬럼명",
        "컬럼설명",
        "영문약어",
        "영문 Full Name",
        "한글단어",
        "오류코드",
        "검증결과",
        "수정방법",
    ]
    worksheet.append(headers)
    for failed in failed_rows:
        worksheet.append(
            [
                failed.source_id,
                failed.schema_name,
                failed.table_name,
                failed.column_name,
                failed.column_description,
                failed.abbreviation,
                failed.full_name,
                failed.korean_word,
                failed.issue_codes,
                failed.validation_messages,
                failed.suggested_actions,
            ]
        )

    _style_header(worksheet, "C00000")
    widths = {
        "A": 14,
        "B": 14,
        "C": 24,
        "D": 26,
        "E": 32,
        "F": 16,
        "G": 30,
        "H": 20,
        "I": 32,
        "J": 52,
        "K": 68,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=cell.column in {5, 9, 10, 11},
            )

    _add_table(
        worksheet,
        display_name="FailedMappings",
        reference=f"A1:K{len(failed_rows) + 1}",
        style_name="TableStyleMedium3",
    )
    worksheet.auto_filter.ref = f"A1:K{len(failed_rows) + 1}"


def _style_header(worksheet, fill_color: str) -> None:
    header_fill = PatternFill("solid", fgColor=fill_color)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28


def _add_table(
    worksheet,
    display_name: str,
    reference: str,
    style_name: str,
) -> None:
    table = Table(displayName=display_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _cell_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _optional_cell_text(
    row: tuple[object, ...], index: int | None
) -> str | None:
    if index is None:
        return None
    value = _cell_text(row, index)
    return value or None
